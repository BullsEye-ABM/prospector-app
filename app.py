# ══════════════════════════════════════════════════════════════════════════════
# PROSPECTOR APP  ·  Pipeline de Prospección B2B  ·  Multi-cliente
# ══════════════════════════════════════════════════════════════════════════════
import streamlit as st
import json, time, re, io, uuid
from datetime import datetime, timedelta
from typing import List, Dict, Optional

st.set_page_config(
    page_title="Prospector App by BullsEye",
    page_icon="🎯",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
  .block-container { padding-top: 2rem; }
  .metric-box {
    background:#F8F9FA; border-radius:10px; padding:16px 12px;
    text-align:center; border:1px solid #E2E8F0;
  }
  .metric-box .val { font-size:2rem; font-weight:700; color:#1F4E79; }
  .metric-box .lbl { font-size:.8rem; color:#64748B; margin-top:2px; }
  .client-card {
    border:1px solid #E2E8F0; border-radius:10px; padding:14px 16px;
    margin-bottom:10px; background:#fff;
  }
  .client-card.active { border-color:#1F4E79; background:#EBF5FB; }
</style>
""", unsafe_allow_html=True)

# ── Auth helpers (definidos aquí, gate se ejecuta más abajo tras SupabaseDB) ──
import hashlib as _hl

def _hash_pw(pw: str) -> str:
    return _hl.sha256(pw.encode("utf-8")).hexdigest()

if "authenticated" not in st.session_state:
    st.session_state.authenticated = False
    st.session_state.auth_name     = ""
    st.session_state.auth_username = ""
    st.session_state.auth_is_admin = False
    st.session_state.auth_token    = ""

# ── Session State ─────────────────────────────────────────────────────────────
for k, v in {
    "selected_client_id": None,
    "selected_client"   : None,
    "clients_list"      : None,   # None = not loaded yet
    "propuesta_de_valor": None,
    "icp": None, "buyer_persona": None, "criterios": None,
    "n_empresas": 20,
    "empresas": [], "empresas_aprobadas": [],
    "contactos_limite": 5,
    "contactos_clay": [], "contactos_aprobados": [], "contactos_final": [],
    "done_pv": False, "done_icp": False, "done_bp": False, "done_empresas": False,
    "done_clay": False, "done_enrich": False,
    "clay_pushed": False, "contacts_pushed_to_enrich": False,
    "show_client_form": False,
    "editing_client_id": None,
    "evaboot_search_id"      : None,
    "evaboot_status"         : None,
    "processed_domains"      : [],
    "processed_contacts"     : [],
    "contactos_aprobacion"   : {},
    "lemlist_campaign_id"    : "",
    "lemlist_campaign_name"  : "",
    "empresas_rechazadas"    : [],
    "_hs_excel_bytes"        : None,
    "_hs_excel_fname"        : "",
    "view_mode"              : "pipeline",
}.items():
    if k not in st.session_state:
        st.session_state[k] = v

# ── API Keys (globales SOi Digital) ───────────────────────────────────────────
def _s(k, default=""):
    try:    return st.secrets.get(k, default)
    except: return default

ANTHROPIC_API_KEY = _s("ANTHROPIC_API_KEY")
LUSHA_API_KEY     = _s("LUSHA_API_KEY")
EVABOOT_API_KEY   = _s("EVABOOT_API_KEY")
SUPABASE_URL      = _s("SUPABASE_URL")
SUPABASE_KEY      = _s("SUPABASE_KEY")
LUSHA_RATE_LIMIT  = 1.0

# Key de Lemlist del cliente activo (se lee de Supabase, no de Secrets)
def lemlist_key():
    c = st.session_state.selected_client
    return (c.get("lemlist_api_key") or "").strip() if c else ""

# ══════════════════════════════════════════════════════════════════════════════
# SUPABASE
# ══════════════════════════════════════════════════════════════════════════════
import requests as _req

class SupabaseDB:
    def __init__(self, url, key):
        self.base = f"{url.rstrip('/')}/rest/v1"
        self.h = {
            "apikey": key,
            "Authorization": f"Bearer {key}",
            "Content-Type": "application/json",
            "Prefer": "return=representation",
        }

    def _get(self, path, params=None):
        r = _req.get(f"{self.base}{path}", headers=self.h, params=params, timeout=10)
        r.raise_for_status(); return r.json()

    def _post(self, path, data):
        r = _req.post(f"{self.base}{path}", headers=self.h, json=data, timeout=10)
        r.raise_for_status(); return r.json()

    def _patch(self, path, data):
        r = _req.patch(f"{self.base}{path}", headers=self.h, json=data, timeout=10)
        r.raise_for_status(); return r.json()

    def _delete(self, path):
        r = _req.delete(f"{self.base}{path}", headers=self.h, timeout=10)
        r.raise_for_status()

    # ── Clientes ──────────────────────────────────────────────────────────────
    def list_clients(self):
        return self._get("/clients", {"order": "name"})

    def create_client(self, data: dict):
        result = self._post("/clients", data)
        return result[0] if result else {}

    def update_client(self, client_id: str, data: dict):
        result = self._patch(f"/clients?id=eq.{client_id}", data)
        return result[0] if result else {}

    def delete_client(self, client_id: str):
        self._delete(f"/clients?id=eq.{client_id}")

    # ── Usuarios ──────────────────────────────────────────────────────────────
    def list_users(self):
        return self._get("/app_users", {"order": "created_at"})

    def get_user(self, username: str):
        rows = self._get("/app_users", {"username": f"eq.{username}", "limit": "1"})
        return rows[0] if rows else None

    def create_user(self, username: str, name: str, password_hash: str, is_admin: bool = False):
        result = self._post("/app_users", {
            "username"     : username,
            "name"         : name,
            "password_hash": password_hash,
            "is_admin"     : is_admin,
        })
        return result[0] if result else {}

    def delete_user(self, username: str):
        self._delete(f"/app_users?username=eq.{username}")

    def count_users(self):
        r = _req.get(f"{self.base}/app_users", headers={**self.h, "Prefer": "count=exact"},
                     params={"select": "id"}, timeout=10)
        return int(r.headers.get("Content-Range","0/0").split("/")[-1])

    # ── Historial ─────────────────────────────────────────────────────────────
    def save_run(self, client_id, empresas, contactos, stats):
        result = self._post("/prospect_runs", {
            "client_id": client_id,
            "empresas" : empresas,
            "contactos": contactos,
            "stats"    : stats,
        })
        return result[0] if result else {}

    def get_runs(self, client_id):
        return self._get("/prospect_runs",
                         {"client_id": f"eq.{client_id}", "order": "run_date.desc"})

    def get_all_runs(self):
        return self._get("/prospect_runs", {"order": "run_date.desc"})

    # ── Sesiones persistentes ──────────────────────────────────────────────────
    def create_session(self, token: str, username: str, name: str, is_admin: bool):
        expires = (datetime.utcnow() + timedelta(days=30)).isoformat()
        try:
            self._post("/app_sessions", {
                "token"     : token,
                "username"  : username,
                "name"      : name,
                "is_admin"  : is_admin,
                "expires_at": expires,
            })
        except Exception:
            pass

    def get_session(self, token: str):
        try:
            rows = self._get("/app_sessions", {
                "token"     : f"eq.{token}",
                "expires_at": f"gt.{datetime.utcnow().isoformat()}",
                "limit"     : "1",
            })
            return rows[0] if rows else None
        except Exception:
            return None

    def delete_session(self, token: str):
        try:
            self._delete(f"/app_sessions?token=eq.{token}")
        except Exception:
            pass

    def log_usage(self, service: str, action: str, units: int = 1,
                  client_id: str = "", client_name: str = "",
                  user_username: str = "", details: dict = None):
        try:
            self._post("/usage_log", {
                "user_username": user_username,
                "client_id"    : client_id,
                "client_name"  : client_name,
                "service"      : service,
                "action"       : action,
                "units"        : units,
                "details"      : details or {},
            })
        except Exception:
            pass  # Non-blocking

    def get_usage_logs(self, limit: int = 1000):
        return self._get("/usage_log", {"order": "created_at.desc", "limit": str(limit)})


@st.cache_resource
def get_db():
    return SupabaseDB(SUPABASE_URL, SUPABASE_KEY) if SUPABASE_URL and SUPABASE_KEY else None

def load_clients():
    db = get_db()
    if db:
        try:
            st.session_state.clients_list = db.list_clients()
        except Exception as e:
            st.session_state.clients_list = []
            st.warning(f"No se pudo conectar a Supabase: {e}")
    else:
        st.session_state.clients_list = []

# ══════════════════════════════════════════════════════════════════════════════
# AUTH — Login gate con sesión persistente (token en URL)
# ══════════════════════════════════════════════════════════════════════════════
_auth_db = SupabaseDB(SUPABASE_URL, SUPABASE_KEY) if SUPABASE_URL and SUPABASE_KEY else None

def _do_login(udata: dict):
    """Autentica en session_state y crea token persistente en query params."""
    token = str(uuid.uuid4())
    st.session_state.authenticated = True
    st.session_state.auth_name     = udata.get("name", "")
    st.session_state.auth_username = udata.get("username", "")
    st.session_state.auth_is_admin = bool(udata.get("is_admin", False))
    st.session_state.auth_token    = token
    if _auth_db:
        _auth_db.create_session(token, udata["username"], udata["name"], bool(udata.get("is_admin", False)))
    st.query_params["t"] = token

# ── Auto-login desde token en URL ─────────────────────────────────────────────
if not st.session_state.authenticated:
    _url_token = st.query_params.get("t", "")
    if _url_token and _auth_db:
        _sess = _auth_db.get_session(_url_token)
        if _sess:
            st.session_state.authenticated = True
            st.session_state.auth_name     = _sess.get("name", "")
            st.session_state.auth_username = _sess.get("username", "")
            st.session_state.auth_is_admin = bool(_sess.get("is_admin", False))
            st.session_state.auth_token    = _url_token

if not st.session_state.authenticated:
    _lc1, _lc2, _lc3 = st.columns([1, 1.1, 1])
    with _lc2:
        st.markdown("<br><br>", unsafe_allow_html=True)
        st.markdown("## 🎯 Prospector App by BullsEye")

        _has_users = False
        if _auth_db:
            try:
                _has_users = _auth_db.count_users() > 0
            except Exception:
                _has_users = False

        if not _has_users:
            # ── Primer uso: crear administrador ──────────────────────────────
            st.markdown("#### ⚙️ Configuración inicial")
            st.info("No hay usuarios registrados. Crea el primer administrador.")
            with st.form("setup_form"):
                _s_name = st.text_input("Nombre completo *", placeholder="Ej: Karmy Cote")
                _s_user = st.text_input("Usuario *", placeholder="Ej: karmy")
                _s_pwd  = st.text_input("Contraseña *", type="password")
                _s_pwd2 = st.text_input("Confirmar contraseña *", type="password")
                _s_ok   = st.form_submit_button("Crear administrador", type="primary", use_container_width=True)
            if _s_ok:
                if not _s_name or not _s_user or not _s_pwd:
                    st.error("Todos los campos son obligatorios.")
                elif _s_pwd != _s_pwd2:
                    st.error("Las contraseñas no coinciden.")
                elif len(_s_pwd) < 6:
                    st.error("La contraseña debe tener al menos 6 caracteres.")
                else:
                    try:
                        _auth_db.create_user(_s_user.strip().lower(), _s_name.strip(), _hash_pw(_s_pwd), is_admin=True)
                        st.success("✅ Administrador creado. Ya puedes iniciar sesión.")
                        st.rerun()
                    except Exception as _e_setup:
                        st.error(f"Error al crear usuario: {_e_setup}")
        else:
            # ── Login normal ─────────────────────────────────────────────────
            st.markdown("#### Iniciar sesión")
            with st.form("login_form"):
                _usr = st.text_input("Usuario", placeholder="tu_usuario")
                _pwd = st.text_input("Contraseña", type="password", placeholder="••••••••")
                _ok  = st.form_submit_button("Ingresar", type="primary", use_container_width=True)
            if _ok:
                try:
                    _udata = _auth_db.get_user(_usr.strip().lower()) if _auth_db else None
                    if _udata and _hash_pw(_pwd) == _udata.get("password_hash", ""):
                        _do_login(_udata)
                        st.rerun()
                    else:
                        st.error("Usuario o contraseña incorrectos.")
                except Exception as _e_login:
                    st.error(f"Error de autenticación: {_e_login}")
    st.stop()

def _parse_json_field(val):
    """Parsea un campo que puede ser dict, list, o JSON string (por si Supabase devuelve texto)."""
    if isinstance(val, (dict, list)):
        return val
    if isinstance(val, str):
        try:
            import json as _json
            return _json.loads(val)
        except Exception:
            return None
    return val

def select_client(client: dict):
    """Carga un cliente y pre-llena ICP/BP en session state."""
    st.session_state.selected_client_id = client["id"]
    st.session_state.selected_client    = client
    # Pre-llenar pipeline con datos guardados del cliente
    # _parse_json_field por si el campo llega como string en vez de dict/list
    st.session_state.propuesta_de_valor = _parse_json_field(client.get("propuesta_de_valor"))
    st.session_state.icp                = _parse_json_field(client.get("icp"))
    st.session_state.buyer_persona      = _parse_json_field(client.get("buyer_persona"))
    st.session_state.criterios          = _parse_json_field(client.get("criterios"))
    st.session_state.done_pv            = bool(client.get("propuesta_de_valor"))
    st.session_state.done_icp           = bool(client.get("icp"))
    st.session_state.done_bp            = bool(client.get("buyer_persona"))
    # Restaurar empresas activas desde Supabase (persisten hasta marcarlas como prospectadas)
    _empresas_guardadas = _parse_json_field(client.get("empresas_activas")) or []
    st.session_state.empresas            = _empresas_guardadas
    st.session_state.empresas_aprobadas  = [e for e in _empresas_guardadas if e.get("aprobada", True)]
    st.session_state.done_empresas       = bool(_empresas_guardadas)
    st.session_state.contactos_clay      = []
    st.session_state.contactos_aprobados = []
    st.session_state.contactos_clay = []
    st.session_state.contactos_final= []
    st.session_state.done_clay      = False
    st.session_state.done_enrich               = False
    st.session_state.clay_pushed               = False
    st.session_state.contacts_pushed_to_enrich = False
    st.session_state.processed_domains    = client.get("processed_domains",  []) or []
    st.session_state.processed_contacts   = client.get("processed_contacts", []) or []
    st.session_state.contactos_aprobacion = {}
    st.session_state.lemlist_campaign_id  = ""
    st.session_state.lemlist_campaign_name= ""
    st.session_state.empresas_rechazadas  = client.get("empresas_rechazadas", []) or []
    # Resetear selector de cargos del Buyer Persona al cambiar de cliente
    if "bp_cargos_sel" in st.session_state:
        del st.session_state["bp_cargos_sel"]

# ══════════════════════════════════════════════════════════════════════════════
# LÓGICA DE NEGOCIO  (igual que antes, sin cambios)
# ══════════════════════════════════════════════════════════════════════════════

# ── 1. AI ─────────────────────────────────────────────────────────────────────
def get_company_recommendations(icp, buyer_persona, criterios, n=20, demo=False, propuesta_de_valor=None, excluir_dominios=None, excluir_nombres=None, razones_rechazo=None, lookalike_empresas=None):
    if demo:
        return [
            {"nombre_empresa":"Alegra",  "dominio_web":"alegra.com",
             "industria":"SaaS B2B","pais":"Colombia","tamano_empleados":"200-300",
             "razon_fit":"SaaS B2B consolidado con equipo comercial activo en LATAM",
             "linkedin_url":"https://www.linkedin.com/company/alegra"},
            {"nombre_empresa":"Conekta", "dominio_web":"conekta.com",
             "industria":"Fintech","pais":"México","tamano_empleados":"100-200",
             "razon_fit":"Fintech B2B con equipo de ventas enterprise",
             "linkedin_url":"https://www.linkedin.com/company/conekta"},
            {"nombre_empresa":"Buk",     "dominio_web":"buk.cl",
             "industria":"SaaS B2B – RRHH","pais":"Chile","tamano_empleados":"300-500",
             "razon_fit":"HR Tech B2B en crecimiento con Series B reciente",
             "linkedin_url":"https://www.linkedin.com/company/buk-cl"},
            {"nombre_empresa":"Siigo",   "dominio_web":"siigo.com",
             "industria":"SaaS B2B","pais":"Colombia","tamano_empleados":"500+",
             "razon_fit":"SaaS para PYMEs con canal de ventas directo",
             "linkedin_url":"https://www.linkedin.com/company/siigo"},
            {"nombre_empresa":"Clip",    "dominio_web":"clip.mx",
             "industria":"Fintech","pais":"México","tamano_empleados":"200-400",
             "razon_fit":"Fintech de pagos B2B con crecimiento sostenido",
             "linkedin_url":"https://www.linkedin.com/company/clip-mx"},
        ]
    import anthropic as _anth
    client = _anth.Anthropic(api_key=ANTHROPIC_API_KEY)
    _pv_context = ""
    if propuesta_de_valor:
        _pv_context = (
            f"Propuesta de valor del vendedor: {propuesta_de_valor.get('propuesta','')}\n"
            f"Dolores que soluciona: {propuesta_de_valor.get('dolores','')}\n\n"
        )
    prompt = (
        "Eres experto en prospección B2B global.\n\n"
        + _pv_context
        + f"ICP:\n{json.dumps(icp, indent=2, ensure_ascii=False)}\n\n"
        f"Buyer Persona:\n{json.dumps(buyer_persona, indent=2, ensure_ascii=False)}\n\n"
        f"Criterios:\n{json.dumps(criterios, indent=2, ensure_ascii=False)}\n\n"
        + (f"Empresas a EXCLUIR por dominio (ya evaluadas): {', '.join((excluir_dominios or [])[:80])}\n\n"
           if excluir_dominios else "")
        + (f"Empresas a EXCLUIR por nombre (ya en cartera del cliente): {', '.join((excluir_nombres or [])[:80])}\n\n"
           if excluir_nombres else "")
        + (f"Razones de rechazo anteriores (evita empresas similares):\n"
           + "\n".join(f"- {rz}" for rz in (razones_rechazo or [])[:20]) + "\n\n"
           if razones_rechazo else "")
        + (f"CLIENTES ACTUALES DEL CLIENTE (empresas de referencia – modelo lookalike):\n"
           + "\n".join(
               f"- {e.get('nombre_empresa','')}"
               + (f" ({e.get('industria','')})" if e.get('industria') else "")
               + (f", {e.get('pais','')}" if e.get('pais') else "")
               + (f", {e.get('tamano_empleados','')}" if e.get('tamano_empleados') else "")
               + (f", {e.get('dominio_web','')}" if e.get('dominio_web') else "")
               for e in (lookalike_empresas or [])[:30]
           ) + "\n"
           "Analiza los patrones comunes de estas empresas (industria, tamaño, país, tipo de negocio, modelo de revenue) "
           "y recomienda empresas con un perfil MUY SIMILAR a ellas. Prioriza empresas que se parezcan a este grupo de referencia.\n\n"
           if lookalike_empresas else "")
        + f"Genera exactamente {n} empresas REALES del mercado objetivo que cumplan todos los criterios.\n"
        "Usa la propuesta de valor y los dolores que soluciona para identificar empresas que realmente necesiten esta solución.\n"
        "Responde ÚNICAMENTE con un JSON array válido. Sin texto adicional.\n"
        "Campos: nombre_empresa, dominio_web, industria, pais, tamano_empleados, razon_fit, linkedin_url"
    )
    resp = _claude_create(client,
        model="claude-sonnet-4-5", max_tokens=4096,
        messages=[{"role":"user","content":prompt}]
    )
    raw = re.sub(r"```(?:json)?\n?","",resp.content[0].text.strip()).strip("`").strip()
    # Intentar parsear; si el JSON está cortado, recuperar objetos completos
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        # Extraer todos los objetos JSON completos del array parcial
        import re as _re
        _matches = _re.findall(r'\{[^{}]+\}', raw, _re.DOTALL)
        _recovered = []
        for _m in _matches:
            try:
                _recovered.append(json.loads(_m))
            except Exception:
                pass
        if _recovered:
            return _recovered
        raise  # si no se recuperó nada, relanzar el error original

# ── Helper: Claude con reintentos para errores 529 (overloaded) ───────────────
def _claude_create(client, **kwargs):
    """Llama a client.messages.create con reintentos para errores 529 (API sobrecargada).
    Si claude-opus-4-6 falla repetidamente, cae a claude-sonnet-4-5 como fallback."""
    import time as _time
    import anthropic as _anth
    # Intentos con modelo principal
    _waits = [3, 8, 20]  # esperas en segundos entre reintentos
    _primary = kwargs.get("model", "")
    _fallback = "claude-sonnet-4-5"  # modelo de respaldo más disponible

    for attempt, wait in enumerate([0] + _waits):
        if wait:
            _time.sleep(wait)
        try:
            return client.messages.create(**kwargs)
        except _anth.APIStatusError as e:
            if e.status_code == 529:
                # Último intento: probar con modelo de respaldo
                if attempt == len(_waits):
                    if _primary != _fallback:
                        _fb_kwargs = {**kwargs, "model": _fallback}
                        return client.messages.create(**_fb_kwargs)
                    raise
                continue
            raise
        except Exception:
            raise

# ── 1a. AI — Análisis de sitio web → Propuesta de Valor ──────────────────────
def analizar_web_empresa(url: str) -> dict:
    """Obtiene el HTML del sitio web y usa Claude para extraer propuesta de valor,
    dolores, cargos sugeridos, roles de compra y pain points del buyer persona."""
    import anthropic as _anth
    # Intentar obtener contenido de la web
    try:
        resp_web = _req.get(url, timeout=10,
            headers={"User-Agent": "Mozilla/5.0 (compatible; ProspectorBot/1.0)"})
        # Extraer solo texto visible, eliminar etiquetas HTML
        html = resp_web.text
        # Limpiar HTML básico
        clean = re.sub(r'<(script|style)[^>]*>.*?</(script|style)>', ' ', html,
                       flags=re.DOTALL | re.IGNORECASE)
        clean = re.sub(r'<[^>]+>', ' ', clean)
        clean = re.sub(r'\s+', ' ', clean).strip()
        contenido = clean[:8000]  # limitar tokens
    except Exception as e:
        contenido = f"[No se pudo obtener el contenido del sitio: {e}]"

    client = _anth.Anthropic(api_key=ANTHROPIC_API_KEY)
    prompt = (
        f"Analiza el siguiente contenido del sitio web de una empresa y extrae información clave.\n\n"
        f"URL: {url}\n"
        f"Contenido del sitio:\n{contenido}\n\n"
        "Devuelve SOLO un JSON válido con esta estructura exacta:\n"
        "{\n"
        '  "propuesta_de_valor": "descripción clara y concisa de qué hace la empresa y qué valor entrega, máximo 300 caracteres",\n'
        '  "dolores_que_soluciona": ["dolor 1", "dolor 2", "dolor 3", ...],\n'
        '  "cargos_sugeridos": ["cargo 1", "cargo 2", ...],\n'
        '  "roles_compra": {\n'
        '    "tomadores_de_decision": ["cargo A", "cargo B"],\n'
        '    "influenciadores": ["cargo C", "cargo D"]\n'
        '  },\n'
        '  "pain_points_buyer": ["pain point 1", "pain point 2", "pain point 3"]\n'
        "}\n\n"
        "Para cargos_sugeridos incluye todos los posibles compradores B2B. "
        "Para dolores_que_soluciona sé específico con los problemas reales del cliente. "
        "Para pain_points_buyer describe los problemas del buyer persona, no de la empresa vendedora."
    )
    resp = _claude_create(client,
        model="claude-opus-4-6", max_tokens=1500,
        messages=[{"role": "user", "content": prompt}]
    )
    raw = re.sub(r"```(?:json)?\n?", "", resp.content[0].text.strip()).strip("`").strip()
    return json.loads(raw)


# ── 1b. AI — Generador de ICP ─────────────────────────────────────────────────
def generar_icp_con_ia(descripcion: str) -> dict:
    """Llama a Claude para generar un ICP estructurado a partir de texto libre."""
    import anthropic as _anth
    client = _anth.Anthropic(api_key=ANTHROPIC_API_KEY)
    prompt = (
        "Eres un experto en estrategia de ventas B2B en Latinoamérica.\n"
        "A partir de la descripción de empresa ideal que te doy, genera un ICP (Ideal Customer Profile) en JSON.\n\n"
        f"Descripción: {descripcion}\n\n"
        "Devuelve SOLO un JSON válido con esta estructura exacta (sin texto adicional):\n"
        "{\n"
        '  "industrias": [lista con 1-4 industrias del catálogo],\n'
        '  "geografias": [lista de países relevantes del mundo],\n'
        '  "tamano_empresa": {"empleados_min": número, "empleados_max": número},\n'
        '  "modelo_negocio": "B2B" | "B2B y B2C" | "Marketplace B2B",\n'
        '  "senales_fit": [lista de señales relevantes],\n'
        '  "exclusiones": [lista de tipos de empresas a excluir]\n'
        "}\n\n"
        "Catálogo de industrias permitidas: SaaS B2B, Fintech, Tecnología / Software, Servicios profesionales, "
        "E-commerce B2B, Edtech, Healthtech, Logística, HR Tech, Ciberseguridad, Inteligencia Artificial / ML, "
        "Marketing Tech, Legal Tech, Proptech / Real Estate, Insurtech, Agritech, Retail / Comercio, "
        "Manufactura, Construcción, Energía / Utilities, Telecomunicaciones, Medios / Entretenimiento, "
        "Viajes / Turismo, Alimentos y Bebidas, Automotriz, Consultoría empresarial, Contabilidad / ERP, "
        "Publicidad / Agencias, BPO / Outsourcing, Salud / Clínicas, Educación corporativa, Gobierno / Sector público.\n"
        "Países: usa los nombres en español de cualquier país del mundo donde aplique el ICP.\n"
        "Señales permitidas (usa solo estas): Tiene equipo de ventas, Tiene equipo de marketing, "
        "Usa CRM, Usa herramientas de sales engagement, Tiene SDRs o BDRs, Está contratando vendedores, "
        "Está contratando roles de revenue, Recibió financiamiento reciente, En expansión a nuevos mercados, "
        "Abrió nuevas oficinas o sedes, Lanzó un nuevo producto recientemente, "
        "Facturación en crecimiento (señales públicas), Tiene página de precios B2B, "
        "Tiene blog o contenido activo, Activo en LinkedIn (publicaciones frecuentes), "
        "Usa publicidad pagada B2B, Tiene caso de éxito / case studies publicados, "
        "Asiste a eventos del sector, Es speaker en conferencias, Miembro de asociaciones del sector, "
        "Usa Salesforce, HubSpot u otro CRM enterprise, Usa Slack o herramientas de productividad cloud, "
        "Tiene API pública o integraciones, Stack tecnológico moderno (SaaS-first)."
    )
    resp = _claude_create(client,
        model="claude-opus-4-6", max_tokens=1024,
        messages=[{"role": "user", "content": prompt}]
    )
    raw = re.sub(r"```(?:json)?\n?", "", resp.content[0].text.strip()).strip("`").strip()
    return json.loads(raw)


# ── 2. Evaboot ────────────────────────────────────────────────────────────────
class EvabootClient:
    BASE = "https://api.evaboot.com/v1"

    def __init__(self, key: str):
        # Evaboot usa "Token" (no "Bearer") según la documentación oficial
        self.h = {"Authorization": f"Token {key}", "Content-Type": "application/json"}

    def submit_extraction(self, linkedin_url: str, search_name: str) -> str:
        """Envía una URL de Sales Navigator a Evaboot. Devuelve search_id."""
        payload = {
            "linkedin_url": linkedin_url,
            "search_name" : search_name,
        }
        r = _req.post(f"{self.BASE}/extractions/url/", headers=self.h, json=payload, timeout=30)
        r.raise_for_status()
        data = r.json()
        # Evaboot devuelve el search_id en el cuerpo (campo 'search_id' o 'id')
        return str(data.get("search_id") or data.get("id") or "")

    def get_extraction_status(self, search_id: str) -> dict:
        """Consulta el estado de una extracción. Devuelve dict con status, progress y prospects."""
        r = _req.get(f"{self.BASE}/extractions/{search_id}/", headers=self.h, timeout=30)
        if r.status_code == 202:
            return {"status": "processing", "progress": 0, "prospects": []}
        r.raise_for_status()
        return r.json()

    def parse_prospects(self, prospects: list) -> list:
        """Convierte los prospects de Evaboot al formato interno.
        Evaboot devuelve campos con Mayúsculas y espacios: 'First Name', 'Current Job', etc."""
        contacts = []
        for p in prospects:
            # Evaboot API devuelve: "First Name", "Last Name", "Current Job",
            # "Company Name", "Company Domain", "Email", "Phone", "Location"
            fn  = (p.get("First Name")  or p.get("first_name")  or p.get("firstName")  or "")
            ln  = (p.get("Last Name")   or p.get("last_name")   or p.get("lastName")   or "")
            full = (p.get("Full Name")  or p.get("full_name")   or p.get("fullName")
                    or p.get("name")    or f"{fn} {ln}".strip())
            c = {
                "first_name"  : fn,
                "last_name"   : ln,
                "full_name"   : full,
                "job_title"   : (p.get("Current Job") or p.get("job_title")
                                 or p.get("headline") or p.get("title")
                                 or p.get("current_title") or ""),
                "company_name": (p.get("Company Name") or p.get("company_name")
                                 or p.get("company") or p.get("current_company") or ""),
                "domain"      : (p.get("Company Domain") or p.get("company_domain")
                                 or p.get("domain") or p.get("company_website") or ""),
                "linkedin_url": (p.get("LinkedIn URL") or p.get("linkedin_url")
                                 or p.get("linkedin_profile_url") or p.get("profileUrl") or ""),
                "country"     : (p.get("Location") or p.get("country") or p.get("location") or ""),
                "email"       : (p.get("Email") or p.get("email") or p.get("found_email") or None),
                "phone"       : (p.get("Phone") or p.get("phone") or p.get("phone_number") or None),
            }
            if c["full_name"] or c["job_title"]:
                contacts.append(c)
        return contacts

    def check_quota(self) -> dict:
        """Verifica créditos disponibles. Devuelve el dict completo de la respuesta."""
        r = _req.get(f"{self.BASE}/quota/", headers=self.h, timeout=15)
        r.raise_for_status()
        return r.json()

    def check_account(self) -> dict:
        """Verifica que la cuenta y Sales Navigator estén correctamente conectados."""
        r = _req.get(f"{self.BASE}/account/", headers=self.h, timeout=15)
        r.raise_for_status()
        return r.json()

# ── Lemlist Client ────────────────────────────────────────────────────────────
# Lemlist usa ?access_token=<key> como query param (Basic auth y x-api-key NO funcionan)
class LemlistClient:
    BASE = "https://api.lemlist.com/api"

    def __init__(self, key: str):
        self.key = key

    def _params(self, extra: dict = None) -> dict:
        """Siempre incluye access_token en los query params."""
        p = {"access_token": self.key}
        if extra:
            p.update(extra)
        return p

    def _get(self, path, params=None):
        r = _req.get(f"{self.BASE}{path}", params=self._params(params), timeout=20)
        if not r.ok:
            try:   detail = r.json()
            except Exception: detail = r.text[:300]
            raise Exception(f"Lemlist {r.status_code}: {detail}")
        return r.json()

    def _post(self, path, payload=None, extra_params=None):
        r = _req.post(f"{self.BASE}{path}", params=self._params(extra_params),
                      json=payload or {}, timeout=20)
        return r.json()

    def _patch(self, path, payload=None):
        r = _req.patch(f"{self.BASE}{path}", params=self._params(),
                       json=payload or {}, timeout=20)
        if not r.ok:
            raise Exception(f"HTTP {r.status_code}: {r.text[:300]}")
        if not r.content or r.status_code == 204:
            return {"status": r.status_code, "ok": True}
        try:
            return r.json()
        except Exception:
            return {"status": r.status_code, "text": r.text[:200]}

    def get_campaigns(self) -> list:
        data = self._get("/campaigns")
        return data if isinstance(data, list) else data.get("campaigns", []) if isinstance(data, dict) else []

    def get_contact_lists(self) -> list:
        """Devuelve las listas de contactos de Lemlist (/contacts/lists)."""
        data = self._get("/contacts/lists")
        return data if isinstance(data, list) else data.get("lists", []) if isinstance(data, dict) else []

    def get_contact_list_leads(self, list_id: str, limit: int = 1000) -> list:
        """Devuelve los leads de una lista de contactos."""
        # Intentar primero /contacts/lists/{id}/leads, luego /contacts?listId=
        try:
            data = self._get(f"/contacts/lists/{list_id}/leads", params={"limit": limit})
            if isinstance(data, list):
                return data
            return data.get("leads", data.get("contacts", []))
        except Exception:
            # Fallback: /contacts?listId={id}
            data = self._get("/contacts", params={"listId": list_id, "limit": limit})
            if isinstance(data, list):
                return data
            return data.get("leads", data.get("contacts", []))

    def create_campaign(self, name: str) -> dict:
        return self._post("/campaigns", {"name": name})

    def add_lead(self, campaign_id: str, lead: dict) -> dict:
        """Agrega un lead a una campaña. email es el identificador único."""
        import urllib.parse as _ulp
        email = (lead.get("email") or
                 f"{lead.get('first_name','x').lower()}.{lead.get('last_name','x').lower()}@noemail.placeholder")
        payload = {k: v for k, v in {
            "firstName"  : lead.get("first_name", ""),
            "lastName"   : lead.get("last_name",  ""),
            "companyName": lead.get("company_name", ""),
            "linkedinUrl": lead.get("linkedin_url", ""),
            "phone"      : lead.get("phone", ""),
            "jobTitle"   : lead.get("job_title", ""),
            "country"    : lead.get("country", ""),
        }.items() if v}
        r = _req.post(
            f"{self.BASE}/campaigns/{campaign_id}/leads/{_ulp.quote(email, safe='')}",
            params=self._params(), json=payload, timeout=15
        )
        return r.json()

    def get_campaign_leads(self, campaign_id: str, limit: int = 500) -> list:
        data = self._get(f"/campaigns/{campaign_id}/leads", params={"limit": limit})
        return data if isinstance(data, list) else data.get("leads", [])

    def get_contacts_by_ids(self, contact_ids: list) -> list:
        """Obtiene detalles completos de contactos dado una lista de contactIds."""
        if not contact_ids:
            return []
        # Lemlist acepta hasta ~50 IDs por request; hacemos batches
        results = []
        batch_size = 50
        for i in range(0, len(contact_ids), batch_size):
            batch = contact_ids[i:i+batch_size]
            try:
                data = self._get("/contacts", params={"idsOrEmails": ",".join(batch)})
                if isinstance(data, list):
                    results.extend(data)
                elif isinstance(data, dict):
                    results.extend(data.get("contacts", data.get("leads", [])))
            except Exception:
                pass
        return results

    def get_campaign_contacts(self, campaign_id: str, limit: int = 500) -> list:
        """Obtiene leads de campaña + sus detalles completos de contacto.
        Preserva el leadId (lea_...) en cada contacto para poder actualizar variables.
        """
        leads = self.get_campaign_leads(campaign_id, limit)
        # Mapa contactId → leadId para recuperarlo después del lookup
        contact_id_to_lead_id = {
            l.get("contactId"): l.get("_id") or l.get("leadId", "")
            for l in leads if l.get("contactId")
        }
        contact_ids = list(contact_id_to_lead_id.keys())
        if not contact_ids:
            return []
        contacts = self.get_contacts_by_ids(contact_ids)
        # Inyectar leadId en cada contacto usando el mapa
        for c in contacts:
            cid = c.get("_id","")
            if cid in contact_id_to_lead_id and not c.get("_leadId"):
                c["_leadId"] = contact_id_to_lead_id[cid]
        return contacts

    def update_lead_variables(self, lead_id: str, variables: dict) -> dict:
        """Crea y setea variables personalizadas (phone1, etc.) en un lead de Lemlist.

        Flujo:
          1. POST /leads/{leadId}/variables  → crea la variable (JSON body, access_token en query)
          2. PATCH /leads/{leadId}/variables → setea valor (Basic Auth + vars como query params)
             Basic Auth evita el conflicto de access_token en query params con las variables.
        """
        _url  = f"{self.BASE}/leads/{lead_id}/variables"
        _hdrs = {"Content-Type": "application/json"}
        # Basic auth: username="", password=api_key (equiv. curl --user ":KEY")
        _basic = ("", self.key)

        def _parse(r):
            if not r.ok:
                raise Exception(f"HTTP {r.status_code}: {r.text[:300]}")
            if not r.content:
                return {"status": r.status_code, "ok": True}
            try:
                return r.json()
            except Exception:
                return {"status": r.status_code, "text": r.text[:200]}

        results = {}
        # Paso 1: crear variable via POST con Basic Auth (sin access_token en query)
        # Lemlist lee TODOS los query params como variables → Basic Auth evita el conflicto
        try:
            r1 = _req.post(_url, auth=_basic, headers=_hdrs, json=variables, timeout=20)
            results["post"] = _parse(r1)
        except Exception as e1:
            results["post_err"] = str(e1)

        # Paso 2: setear valor via PATCH con Basic Auth + vars como query params
        try:
            r2 = _req.patch(_url, auth=_basic, params=variables, timeout=20)
            results["patch"] = _parse(r2)
        except Exception as e2:
            results["patch_err"] = str(e2)

        return results

    def update_campaign_lead_phone(self, campaign_id: str, lead_id: str, phone: str) -> dict:
        """Actualiza el campo estándar 'phone' de un lead via el endpoint de campaña.
        PATCH /campaigns/{campaignId}/leads/{leadId} con body {"phone": "..."}
        Este es el endpoint que controla la columna 'Phone number' en la UI de Lemlist.
        """
        _url  = f"{self.BASE}/campaigns/{campaign_id}/leads/{lead_id}"
        _auth = {"access_token": self.key}
        _hdrs = {"Content-Type": "application/json"}
        r = _req.patch(_url, params=_auth, headers=_hdrs, json={"phone": phone}, timeout=20)
        if not r.ok:
            raise Exception(f"HTTP {r.status_code}: {r.text[:300]}")
        if not r.content:
            return {"status": r.status_code, "ok": True}
        try:
            return r.json()
        except Exception:
            return {"status": r.status_code, "text": r.text[:200]}

    def update_lead(self, email: str, fields: dict) -> dict:
        """Fallback: PATCH /leads/{email} para campos estándar."""
        import urllib.parse as _ulp
        _enc = _ulp.quote(email, safe='')
        return self._patch(f"/leads/{_enc}", fields)

# ── Sales Navigator URL Generator ─────────────────────────────────────────────
_BATCH_SIZE = 15   # máx empresas por URL de Sales Navigator

def generar_url_sales_navigator(buyer_persona: dict, icp: dict,
                                 empresas_aprobadas: list = None,
                                 _company_offset: int = 0) -> str:
    """Genera URL de LinkedIn Sales Navigator con Buyer Persona + ICP + empresas aprobadas."""

    _GEO_IDS = {
        # ── Latinoamérica (IDs verificados por usuario) ───────────────────────
        "Chile"     : "104621616",
        "Perú"      : "102927786", "Peru"       : "102927786",
        "Ecuador"   : "106373116",
        "Colombia"  : "100876405",
        "Argentina" : "100446943",
        "Uruguay"   : "100867946",
        "México"    : "103323778", "Mexico"     : "103323778",
        "Brasil"    : "106057199", "Brazil"     : "106057199",
        "Costa Rica": "101739942",
        "Bolivia"   : "104720398",
        "Paraguay"  : "107721998",
        "Venezuela" : "101490751",
        "Panamá"    : "100808254", "Panama"     : "100808254",
        "Guatemala" : "100999417",
        "Honduras"  : "101517120",
        "El Salvador": "101779815",
        "Nicaragua" : "106522019",
        "Cuba"      : "104512186",
        "República Dominicana": "104604687",
        "Puerto Rico": "101183737",
        # ── Norteamérica (IDs verificados por usuario) ────────────────────────
        "Estados Unidos": "103644278", "United States": "103644278",
        "Canadá"        : "101174742", "Canada"        : "101174742",
        # ── Europa Occidental (IDs verificados por usuario) ───────────────────
        "España"      : "105646813", "Spain"          : "105646813",
        "Reino Unido" : "101165590", "United Kingdom" : "101165590",
        "Alemania": "101282230", "Germany": "101282230",
        "Francia": "105015875", "France": "105015875",
        "Italia": "103350119", "Italy": "103350119",
        "Portugal": "100364837",
        "Países Bajos": "102890719", "Netherlands": "102890719",
        "Bélgica": "100565514", "Belgium": "100565514",
        "Suecia": "105117694", "Sweden": "105117694",
        "Noruega": "103819153", "Norway": "103819153",
        "Dinamarca": "104514075", "Denmark": "104514075",
        "Finlandia": "100456013", "Finland": "100456013",
        "Suiza": "106693272", "Switzerland": "106693272",
        "Austria": "103883259",
        "Irlanda": "104738515", "Ireland": "104738515",
        "Grecia": "104677530", "Greece": "104677530",
        # Europa del Este
        "Polonia": "105072130", "Poland": "105072130",
        "Rumania": "106670623", "Romania": "106670623",
        "República Checa": "104508036", "Czech Republic": "104508036",
        "Hungría": "100288700", "Hungary": "100288700",
        # Oriente Medio
        "Israel": "101620260",
        "Emiratos Árabes Unidos": "104305776", "UAE": "104305776",
        "Arabia Saudita": "109035765", "Saudi Arabia": "109035765",
        "Qatar": "104774139",
        # Asia-Pacífico
        "India": "102713980",
        "China": "102890883",
        "Japón": "101355337", "Japan": "101355337",
        "Corea del Sur": "105149290", "South Korea": "105149290",
        "Singapur": "102454443", "Singapore": "102454443",
        "Australia": "101452733",
        "Nueva Zelanda": "105490917", "New Zealand": "105490917",
        "Indonesia": "102478259",
        "Malasia": "102905671", "Malaysia": "102905671",
        "Vietnam": "104195383",
        "Tailandia": "102996343", "Thailand": "102996343",
        "Filipinas": "103121230", "Philippines": "103121230",
        # África
        "Sudáfrica": "104035573", "South Africa": "104035573",
        "Egipto": "106155005", "Egypt": "106155005",
        "Nigeria": "105365761",
    }
    _HEADCOUNT_MAP = {
        "1-10": "A", "11-50": "B", "51-200": "C", "201-500": "D",
        "501-1.000": "E", "501-1000": "E", "1.001-5.000": "F",
        "1001-5000": "F", "5.001-10.000": "G", "5001-10000": "G",
        "10.000+": "H", "10000+": "H", "+10.000": "H",
    }

    import urllib.parse as _ul

    def _snav(text: str) -> str:
        """Doble-encode para valores dentro del query de Sales Navigator.
        Espacios → %2520, acentos → %25C3%25XX, etc."""
        return _ul.quote(str(text).strip(), safe="").replace("%", "%25")

    titulos         = (buyer_persona or {}).get("cargos_objetivo",  [])
    titulos_excluir = (buyer_persona or {}).get("cargos_excluidos", [])

    filter_parts = []

    # ── 1. Empresas aprobadas (CURRENT_COMPANY) ───────────────────────────────
    if empresas_aprobadas:
        company_names = [e.get("nombre_empresa", "") for e in empresas_aprobadas if e.get("nombre_empresa")]
        # Usar el batch solicitado (para dividir listas grandes en varias URLs)
        _batch = company_names[_company_offset : _company_offset + _BATCH_SIZE]
        if _batch:
            comp_vals = "%2C".join(
                f"(text%3A{_snav(n)}%2CselectionType%3AINCLUDED)"
                for n in _batch
            )
            filter_parts.append(f"(type%3ACURRENT_COMPANY%2Cvalues%3AList({comp_vals}))")

    # ── 2. Geografía: GEOGRAPHY + REGION (mismo ID, ambos requeridos) ─────────
    _geo_val = (icp or {}).get("geografias", [])
    paises   = (icp or {}).get("paises", []) or \
               (_geo_val if isinstance(_geo_val, list) else [])
    _manual_geo_id = (icp or {}).get("linkedin_geo_id", "").strip()

    geo_entries = []
    if _manual_geo_id and paises:
        geo_entries.append((paises[0], _manual_geo_id))
        for p in paises[1:4]:
            gid = _GEO_IDS.get(p)
            if gid:
                geo_entries.append((p, gid))
    else:
        for p in paises[:4]:
            gid = _GEO_IDS.get(p)
            if gid:
                geo_entries.append((p, gid))

    if geo_entries:
        geo_vals = "%2C".join(
            f"(id%3A{gid}%2Ctext%3A{_snav(p)}%2CselectionType%3AINCLUDED)"
            for p, gid in geo_entries
        )
        filter_parts.append(f"(type%3AGEOGRAPHY%2Cvalues%3AList({geo_vals}))")
        filter_parts.append(f"(type%3AREGION%2Cvalues%3AList({geo_vals}))")

    # ── 3. Exclusión estricta de cargos (CURRENT_JOB_TITLE EXCLUDED) ──────────
    if titulos_excluir:
        excl_vals = "%2C".join(
            f"(text%3A{_snav(e)}%2CselectionType%3AEXCLUDED)"
            for e in titulos_excluir if e.strip()
        )
        if excl_vals:
            # Sales Navigator requiere AMBOS filtros para que la exclusión funcione
            filter_parts.append(f"(type%3ACURRENT_JOB_TITLE%2Cvalues%3AList({excl_vals}))")
            filter_parts.append(f"(type%3ACURRENT_TITLE%2Cvalues%3AList({excl_vals}))")

    # ── 4. Construir query final: filters primero, keywords después ────────────
    # Formato correcto verificado con URLs manuales de Sales Navigator:
    # ?query=(filters%3AList(...)%2Ckeywords%3AKEYWORD1%2520OR%2520KEYWORD2)
    query_parts = []

    if filter_parts:
        query_parts.append(f"filters%3AList({('%2C').join(filter_parts)})")

    if titulos:
        # Keywords con doble-encoding (espacios → %2520, acentos → %25XX)
        # Sin límite de cargos — Sales Navigator soporta URLs largas
        kw_str = " OR ".join(
            t.strip().replace('"', '').replace("(", "").replace(")", "")
            for t in titulos
        )
        query_parts.append(f"keywords%3A{_snav(kw_str)}")

    if query_parts:
        return f"https://www.linkedin.com/sales/search/people?query=({'%2C'.join(query_parts)})"
    return "https://www.linkedin.com/sales/search/people"


# ── Google Sheets ──────────────────────────────────────────────────────────────
def _get_gsheet_client():
    """Devuelve un cliente gspread autenticado con service account."""
    import gspread
    from google.oauth2.service_account import Credentials
    creds_raw = _s("GSHEETS_CREDENTIALS")
    if not creds_raw:
        return None
    try:
        creds_dict = json.loads(creds_raw) if isinstance(creds_raw, str) else dict(creds_raw)
        scopes = ["https://spreadsheets.google.com/feeds",
                  "https://www.googleapis.com/auth/drive"]
        creds = Credentials.from_service_account_info(creds_dict, scopes=scopes)
        return gspread.authorize(creds)
    except Exception as e:
        st.error(f"Error autenticando Google Sheets: {e}")
        return None

def write_to_gsheets(sheet_url: str, empresas: list, contactos: list, client_name: str = ""):
    """Escribe empresas y contactos en dos pestañas del Google Sheet."""
    gc = _get_gsheet_client()
    if not gc:
        raise ValueError("No hay credenciales de Google Sheets configuradas (GSHEETS_CREDENTIALS).")

    sh = gc.open_by_url(sheet_url)
    today = datetime.now().strftime("%Y-%m-%d %H:%M")

    # ── Pestaña Empresas ──────────────────────────────────────────────────────
    try:
        ws_emp = sh.worksheet("Empresas")
        ws_emp.clear()
    except Exception:
        ws_emp = sh.add_worksheet("Empresas", rows=1000, cols=10)

    emp_headers = ["Empresa","Dominio","País","Industria","Tamaño","Razón de Fit","LinkedIn","Fecha"]
    emp_rows = [[
        e.get("nombre_empresa",""), e.get("dominio_web",""),
        e.get("pais",""),           e.get("industria",""),
        e.get("tamano_empleados",""), e.get("razon_fit",""),
        e.get("linkedin_url",""),   today,
    ] for e in empresas]
    ws_emp.update([emp_headers] + emp_rows)

    # ── Pestaña Contactos ─────────────────────────────────────────────────────
    try:
        ws_con = sh.worksheet("Contactos")
        ws_con.clear()
    except Exception:
        ws_con = sh.add_worksheet("Contactos", rows=5000, cols=15)

    con_headers = ["Nombre","Empresa","Cargo","Rol de compra","Email","Fuente Email",
                   "Teléfono","Fuente Teléfono","LinkedIn","País","Industria","Fecha","Cliente"]
    con_rows = [[
        norm_person(c.get("full_name",""), c.get("first_name",""), c.get("last_name","")),
        norm_company(c.get("company_name","")),
        c.get("job_title",""),
        c.get("rol_compra",""),
        c.get("email",""),
        c.get("email_source",""),
        c.get("phone",""),
        c.get("phone_source",""),
        c.get("linkedin_url",""),
        c.get("country",""),
        c.get("industry",""),
        today, client_name,
    ] for c in contactos]
    ws_con.update([con_headers] + con_rows)

    return len(emp_rows), len(con_rows)


_TITULOS_DECISION = {
    "ceo","co-ceo","founder","co-founder","owner","president","managing director",
    "cro","chief revenue officer","coo","cfo","cto","cio","chro","chief people officer",
    "vp of sales","vp comercial","vp revenue","vp of marketing","vp of operations",
    "vp of engineering","vp of product","vp of finance",
    "director comercial","director de ventas","director de revenue","director de marketing",
    "director de operaciones","sales director","finance director","it director",
    "general manager","country manager","executive director",
}

def clasificar_rol_compra(job_title: str, roles_ia: dict = None) -> str:
    """Clasifica un contacto como Tomador de decisión o Influenciador."""
    t = (job_title or "").lower().strip()
    # Primero verificar contra las listas generadas por la IA
    if roles_ia:
        for cargo in roles_ia.get("tomadores_de_decision", []):
            if cargo.lower() in t or t in cargo.lower():
                return "Tomador de decisión"
        for cargo in roles_ia.get("influenciadores", []):
            if cargo.lower() in t or t in cargo.lower():
                return "Influenciador"
    # Fallback: reglas por título
    for patron in _TITULOS_DECISION:
        if patron in t:
            return "Tomador de decisión"
    return "Influenciador"


def _demo_contacts():
    return [
        {"first_name":"Carlos",  "last_name":"Garcia Lopez",   "full_name":"Carlos Garcia Lopez",
         "job_title":"Director Comercial","company_name":"Alegra", "domain":"alegra.com",
         "email":"carlos.garcia@alegra.com","email_source":"Clay",
         "linkedin_url":"https://linkedin.com/in/carlosgarcia","country":"Colombia","industry":"SaaS B2B","phone":None},
        {"first_name":"Maria",   "last_name":"Rodriguez Perez","full_name":"Maria Rodriguez Perez",
         "job_title":"VP of Sales","company_name":"Conekta","domain":"conekta.com",
         "email":None,"email_source":None,
         "linkedin_url":"https://linkedin.com/in/mariarodriguez","country":"México","industry":"Fintech","phone":None},
        {"first_name":"Andres",  "last_name":"Martinez Silva", "full_name":"Andres Martinez Silva",
         "job_title":"CEO","company_name":"Buk","domain":"buk.cl",
         "email":"andres@buk.cl","email_source":"Clay",
         "linkedin_url":"https://linkedin.com/in/andresmartinez","country":"Chile","industry":"SaaS B2B","phone":None},
        {"first_name":"Patricia","last_name":"Vargas Mora",    "full_name":"Patricia Vargas Mora",
         "job_title":"Head of Sales","company_name":"Siigo","domain":"siigo.com",
         "email":"p.vargas@siigo.com","email_source":"Clay",
         "linkedin_url":"https://linkedin.com/in/patriciavargas","country":"Colombia","industry":"SaaS B2B","phone":None},
        {"first_name":"Roberto", "last_name":"Fuentes Mora",   "full_name":"Roberto Fuentes Mora",
         "job_title":"Founder","company_name":"Clip","domain":"clip.mx",
         "email":None,"email_source":None,
         "linkedin_url":"https://linkedin.com/in/robertofuentes","country":"México","industry":"Fintech","phone":None},
    ]

# ── 3. Lusha ──────────────────────────────────────────────────────────────────
def _lusha_person(fn, ln, company, li=None, key=None, demo=False):
    """Busca teléfono Y email en Lusha V2. Prioriza LinkedIn URL si está disponible.
    Endpoint: GET https://api.lusha.com/v2/person (V1 deprecado 14-ene-2025).
    """
    if demo:
        import random
        px=["+52 55","+57 300","+56 9","+54 11","+57 310"]
        phone = f"{random.choice(px)} {random.randint(10000000,99999999)}" if random.random()>.35 else None
        f=fn.lower().replace(" ",""); l=ln.lower().split()[0] if ln else "x"
        domain=(company or "").lower().replace(" ","")+".com"
        email_demo = f"{f}.{l}@{domain}" if random.random()>.55 else None
        return phone, email_demo
    if not key: return None, None

    _BASE = "https://api.lusha.com/v2/person"
    _HDR  = {"api_key": key}

    def _parse_lusha_response(r):
        if r.status_code == 200:
            raw = r.json()
            # V2 estructura: {"contact": {"data": {...} | null, "isCreditCharged": bool, "error": {...}}}
            d = raw.get("contact", raw)   # desenvuelve clave "contact" si existe
            if isinstance(d, dict) and "data" in d:
                d = d["data"] or {}       # puede ser null cuando EMPTY_DATA
            if not d:
                return None, None         # contacto no encontrado en base Lusha
            phones = d.get("phoneNumbers", [])
            direct = [ph for ph in phones if ph.get("type") in ("direct","mobile","work")]
            phone  = direct[0].get("number") if direct else (phones[0].get("number") if phones else None)
            emails = d.get("emails", [])
            email  = emails[0].get("email") if emails else None
            return phone, email
        elif r.status_code == 402:
            st.warning("⚠️ Créditos Lusha agotados")
        elif r.status_code == 404:
            pass  # endpoint no encontrado
        return None, None

    try:
        # Intento 1: solo LinkedIn URL (más preciso)
        if li and li.strip():
            r = _req.get(_BASE,
                         headers=_HDR,
                         params={"linkedinUrl": li.strip()},
                         timeout=10)
            phone, email = _parse_lusha_response(r)
            if phone or email:
                return phone, email

        # Intento 2: nombre + apellido + empresa
        p = {"firstName": fn, "lastName": ln, "companyName": company}
        r = _req.get(_BASE, headers=_HDR, params=p, timeout=10)
        return _parse_lusha_response(r)

    except Exception as e:
        st.warning(f"Lusha: {e}")
    return None, None

# ── 4. Lemlist ────────────────────────────────────────────────────────────────
def _lemlist_one(fn, ln, domain, key=None, demo=False):
    if demo:
        import random
        f=fn.lower().replace(" ",""); l=ln.lower().split()[0] if ln else ""
        opts=[f"{f}.{l}@{domain}",f"{f[0]}{l}@{domain}",f"{f}@{domain}"]
        return random.choice(opts) if random.random()>.45 else None
    if not key or not domain: return None
    try:
        r=_req.get("https://api.lemlist.com/api/leads/email-finder",
                   auth=(key,""),params={"firstName":fn,"lastName":ln,"domain":domain},timeout=10)
        if r.status_code==200:
            d=r.json()
            if d.get("email") and float(d.get("confidence",0))>=0.70: return d["email"]
    except Exception as e: st.warning(f"Lemlist: {e}")
    return None

# ── 5. Waterfall de enriquecimiento ───────────────────────────────────────────
def enrich_waterfall(contacts, lusha_key=None, lemlist_key_val=None,
                     demo=False, pb=None):
    """
    Teléfono:  Lusha
    Email:     Lemlist -> Lusha
    """
    result = []
    total  = len(contacts)

    for i, c in enumerate(contacts):
        n = c.copy()
        fn = n.get("first_name",""); ln = n.get("last_name","")
        company = n.get("company_name",""); li = n.get("linkedin_url","")
        domain  = n.get("domain","")

        # ── TELÉFONO: Lusha ────────────────────────────────────────────────────
        if not n.get("phone"):
            phone, lusha_email = _lusha_person(fn, ln, company, li,
                                               key=lusha_key, demo=demo)
            if phone:
                n["phone"] = phone; n["phone_source"] = "Lusha"
                if not n.get("email") and lusha_email:
                    n["email"] = lusha_email; n["email_source"] = "Lusha"

        # ── EMAIL: Lemlist → Lusha ─────────────────────────────────────────────
        if not n.get("email"):
            # 1. Lemlist
            em = _lemlist_one(fn, ln, domain, key=lemlist_key_val, demo=demo)
            if em:
                n["email"] = em; n["email_source"] = "Lemlist"
            # 2. Lusha email fallback
            if not n.get("email") and not demo:
                _, lusha_em = _lusha_person(fn, ln, company, li, key=lusha_key, demo=False)
                if lusha_em:
                    n["email"] = lusha_em; n["email_source"] = "Lusha"

        result.append(n)
        if pb: pb.progress((i+1)/total, text=f"Enriqueciendo {i+1}/{total}: {n.get('full_name','')}")
        time.sleep(LUSHA_RATE_LIMIT)

    return result

# ── 5. Normalización ──────────────────────────────────────────────────────────
_PART={"de","la","del","los","las","el","y","e","von","van","da","di"}
_SUF=[r"\bS\.?A\.?S\.?\b",r"\bS\.?A\.?\b",r"\bInc\.?\b",r"\bLLC\.?\b",r"\bLtd\.?\b",
      r"\bGmbH\b",r"\bS\.?R\.?L\.?\b",r"\bS\.?de R\.?L\.?\b",
      r"\bS\.?de R\.?L\.? de C\.?V\.?\b",r"\bS\.?A\.? de C\.?V\.?\b"]

def _cap(t):
    w,o=t.lower().split(),[]
    for i,x in enumerate(w):
        if x in _PART and i>0: o.append(x)
        elif x.upper()==x and 1<len(x)<=5: o.append(x.upper())
        else: o.append(x.capitalize())
    return " ".join(o)

def norm_person(full="",first="",last=""):
    if first and last:
        fn=first.strip().split()[0] if first.strip() else ""
        ln=last.strip().split()[0]  if last.strip()  else ""
        name=f"{fn} {ln}".strip() if fn or ln else full
    elif full:
        p=full.strip().split()
        name=f"{p[0]} {p[1]}" if len(p)>=2 else p[0] if p else ""
    else: name=first or ""
    return _cap(name)

def norm_company(company):
    r=company.strip()
    for s in _SUF: r=re.sub(s,"",r,flags=re.IGNORECASE).strip()
    r=r.strip(",.").strip()
    if r.isupper() and len(r)<=6: return r
    return _cap(r)

# ── 6. Excel ──────────────────────────────────────────────────────────────────
def build_excel(contacts, client_name="") -> bytes:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import pandas as pd
    today=datetime.now().strftime("%Y-%m-%d")
    rows=[{
        "Nombre"               :norm_person(c.get("full_name",""),c.get("first_name",""),c.get("last_name","")),
        "Empresa"              :norm_company(c.get("company_name","")),
        "Cargo"                :c.get("job_title",""),
        "Email"                :c.get("email",""),
        "Teléfono"             :c.get("phone",""),
        "LinkedIn"             :c.get("linkedin_url",""),
        "País"                 :c.get("country",""),
        "Industria"            :c.get("industry",c.get("industria","")),
        "Fuente Email"         :c.get("email_source",""),
        "Fuente Teléfono"      :c.get("phone_source",""),
        "Fecha Enriquecimiento":today,
        "Cliente"              :client_name,
    } for c in contacts]
    df=pd.DataFrame(rows)
    buf=io.BytesIO()
    with pd.ExcelWriter(buf,engine="openpyxl") as writer:
        df.to_excel(writer,sheet_name="Prospectos",index=False)
        wb=writer.book; ws=writer.sheets["Prospectos"]
        thin=Side(style="thin",color="CCCCCC")
        border=Border(left=thin,right=thin,top=thin,bottom=thin)
        h_font=Font(bold=True,color="FFFFFF",size=11)
        h_fill=PatternFill(start_color="1F4E79",end_color="1F4E79",fill_type="solid")
        h_align=Alignment(horizontal="center",vertical="center",wrap_text=True)
        alt=PatternFill(start_color="EBF5FB",end_color="EBF5FB",fill_type="solid")
        for cell in ws[1]:
            cell.font=h_font; cell.fill=h_fill
            cell.alignment=h_align; cell.border=border
        ws.row_dimensions[1].height=35
        for ri in range(2,len(df)+2):
            for cell in ws[ri]:
                if ri%2==0: cell.fill=alt
                cell.border=border; cell.alignment=Alignment(vertical="center")
            ws.row_dimensions[ri].height=20
        for col,w in zip("ABCDEFGHIJKL",[22,22,28,32,20,38,12,22,14,16,18,20]):
            ws.column_dimensions[col].width=w
        ws.freeze_panes="A2"
        # Resumen
        total=len(df); ce=int((df["Email"]!="").sum()); ct=int((df["Teléfono"]!="").sum())
        sc=int((df["Fuente Email"]=="Clay").sum()); sl=int((df["Fuente Email"]=="Lemlist").sum())
        ws2=wb.create_sheet("Resumen")
        ws2.column_dimensions["A"].width=30; ws2.column_dimensions["B"].width=18
        for row in [["RESUMEN",""],["Cliente",client_name],["Fecha",today],["",""],
                    ["Total prospectos",total],["Con email",ce],
                    ["  - Clay",sc],["  - Lemlist",sl],["Sin email",total-ce],
                    ["Con teléfono (Lusha)",ct],["Sin teléfono",total-ct],
                    ["Cobertura email",f"{ce/total*100:.0f}%" if total else "0%"],
                    ["Cobertura teléfono",f"{ct/total*100:.0f}%" if total else "0%"]]:
            ws2.append(row)
        ws2["A1"].font=Font(bold=True,size=14)
    return buf.getvalue()


def build_excel_hubspot(contacts, empresas=None, client_name="", particularidades="") -> bytes:
    """Excel listo para importar en HubSpot — sheet Contactos + Empresas + Resumen."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import pandas as pd

    today = datetime.now().strftime("%Y-%m-%d")

    # Lookup empresa por nombre (para enriquecer contactos con datos de empresa)
    empresa_map = {}
    for e in (empresas or []):
        key = (e.get("nombre_empresa") or "").strip().lower()
        if key:
            empresa_map[key] = e

    # ── Columnas adicionales por particularidades del cliente (IA) ────────────
    custom_cols_data = {}  # {contact_email: {col_name: value}}
    if particularidades and particularidades.strip() and contacts:
        try:
            import anthropic as _ant_exc
            _ant_exc_client = _ant_exc.Anthropic()
            _compact_contacts = [
                {"email": c.get("email",""), "cargo": c.get("job_title",""),
                 "pais": c.get("country",""), "empresa": c.get("company_name","")}
                for c in contacts
            ]
            _exc_prompt = (
                f"Eres un asistente de prospección B2B. Para el cliente '{client_name}' necesitas agregar columnas personalizadas al Excel.\n\n"
                f"Instrucciones del cliente:\n{particularidades}\n\n"
                f"Lista de contactos:\n{json.dumps(_compact_contacts, ensure_ascii=False)}\n\n"
                "Para cada contacto, genera los valores de las columnas adicionales según las instrucciones.\n"
                "Responde ÚNICAMENTE con un JSON array (mismo orden y cantidad que la lista). "
                "Cada objeto debe tener solo las columnas adicionales a agregar. "
                "Si no puedes determinar un valor, usa string vacío \"\".\n"
                "Ejemplo de respuesta: [{\"Área\": \"Ventas\"}, {\"Área\": \"Marketing\"}, ...]"
            )
            _exc_resp = _ant_exc_client.messages.create(
                model="claude-opus-4-5",
                max_tokens=4096,
                messages=[{"role": "user", "content": _exc_prompt}],
            )
            _exc_text = _exc_resp.content[0].text.strip()
            if _exc_text.startswith("```"):
                _exc_text = "\n".join(_exc_text.split("\n")[1:-1])
            _exc_list = json.loads(_exc_text)
            for i, c in enumerate(contacts):
                if i < len(_exc_list):
                    custom_cols_data[c.get("email","") or str(i)] = _exc_list[i]
        except Exception:
            pass  # Si falla la IA, continuar sin columnas extra

    # ── Sheet 1: Contactos (columnas exactas que HubSpot acepta en import) ──
    contact_rows = []
    for c in contacts:
        emp_key = (c.get("company_name") or "").strip().lower()
        emp     = empresa_map.get(emp_key, {})
        # Teléfono unificado: usa lo que haya (Lemlist o Lusha), 1 solo campo
        phone_val = c.get("phone","").strip()
        row = {
            "First Name"          : c.get("first_name",""),
            "Last Name"           : c.get("last_name",""),
            "Email Address"       : c.get("email",""),
            "Phone Number"        : phone_val,
            "Job Title"           : c.get("job_title",""),
            "LinkedIn Profile URL": c.get("linkedin_url",""),
            "Company Name"        : norm_company(c.get("company_name","")),
            "Country/Region"      : c.get("country","") or emp.get("pais",""),
            "Website"             : emp.get("dominio_web","") or c.get("domain",""),
            "Industry"            : emp.get("industria","") or c.get("industry",c.get("industria","")),
            "Fuente Email"        : c.get("email_source",""),
            "Fuente Teléfono"     : c.get("phone_source",""),
        }
        # Merge custom columns from particularidades
        _extra = custom_cols_data.get(c.get("email","") or "", {})
        row.update(_extra)
        contact_rows.append(row)

    # ── Sheet 2: Empresas (columnas HubSpot companies import) ──
    seen_companies = set()
    company_rows = []
    for e in (empresas or []):
        name = (e.get("nombre_empresa") or "").strip()
        if not name or name.lower() in seen_companies:
            continue
        seen_companies.add(name.lower())
        company_rows.append({
            "Company name"        : norm_company(name),
            "Domain Name"         : e.get("dominio_web",""),
            "Industry"            : e.get("industria",""),
            "Country/Region"      : e.get("pais",""),
            "Number of Employees" : e.get("tamano_empleados",""),
            "LinkedIn Company Page": e.get("linkedin_url",""),
            "Razón de fit"        : e.get("razon_fit",""),
        })

    df_c = pd.DataFrame(contact_rows)
    df_e = pd.DataFrame(company_rows) if company_rows else pd.DataFrame()

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_c.to_excel(writer, sheet_name="Contactos", index=False)
        if not df_e.empty:
            df_e.to_excel(writer, sheet_name="Empresas", index=False)

        wb = writer.book

        thin   = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        def _style_ws(ws, nrows, header_color):
            h_font  = Font(bold=True, color="FFFFFF", size=11)
            h_fill  = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
            h_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            alt     = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
            for cell in ws[1]:
                cell.font=h_font; cell.fill=h_fill
                cell.alignment=h_align; cell.border=border
            ws.row_dimensions[1].height = 35
            for ri in range(2, nrows+2):
                for cell in ws[ri]:
                    if ri % 2 == 0: cell.fill = alt
                    cell.border=border
                    cell.alignment=Alignment(vertical="center")
                ws.row_dimensions[ri].height = 20
            for col_cells in ws.columns:
                max_len = max((len(str(cell.value or "")) for cell in col_cells), default=10)
                ws.column_dimensions[col_cells[0].column_letter].width = min(max(max_len+2, 12), 50)
            ws.freeze_panes = "A2"

        _style_ws(writer.sheets["Contactos"], len(df_c), "1F4E79")
        if not df_e.empty:
            _style_ws(writer.sheets["Empresas"], len(df_e), "0F7B6C")

        # ── Resumen ──
        ws_r = wb.create_sheet("Resumen")
        ws_r.column_dimensions["A"].width = 32
        ws_r.column_dimensions["B"].width = 18
        total_c   = len(contact_rows)
        con_em    = sum(1 for r in contact_rows if r.get("Email Address",""))
        con_tel   = sum(1 for r in contact_rows if r.get("Phone Number",""))
        con_lusha = sum(1 for c in contacts if c.get("phone_source") == "Lusha")
        for row in [
            ["RESUMEN — Exportación HubSpot", ""],
            ["Cliente",            client_name],
            ["Fecha",              today],
            ["", ""],
            ["Total contactos",    total_c],
            ["  Con email",        con_em],
            ["  Con teléfono",     con_tel],
            ["    → desde Lemlist", con_tel - con_lusha],
            ["    → desde Lusha",   con_lusha],
            ["Sin email",          total_c - con_em],
            ["Sin teléfono",       total_c - con_tel],
            ["", ""],
            ["Cobertura email",    f"{con_em/total_c*100:.0f}%" if total_c else "0%"],
            ["Cobertura teléfono", f"{con_tel/total_c*100:.0f}%" if total_c else "0%"],
            ["", ""],
            ["Total empresas",     len(company_rows)],
        ]:
            ws_r.append(row)
        ws_r["A1"].font = Font(bold=True, size=14)

    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# CARGAR CLIENTES AL INICIO
# ══════════════════════════════════════════════════════════════════════════════
if st.session_state.clients_list is None:
    load_clients()

# Auto-restaurar cliente desde query params al refrescar la página
_qp_client_id = st.query_params.get("c", "")
if (_qp_client_id
        and not st.session_state.selected_client_id
        and st.session_state.clients_list):
    _auto_client = next(
        (c for c in st.session_state.clients_list if str(c["id"]) == str(_qp_client_id)),
        None
    )
    if _auto_client:
        select_client(_auto_client)

# ══════════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("## 🎯 Prospector App by BullsEye")
    st.caption("Pipeline de Prospección B2B")
    st.divider()

    # ── Usuario activo + logout ───────────────────────────────────────────────
    _sb_col1, _sb_col2 = st.columns([3, 1])
    with _sb_col1:
        st.caption(f"👤 **{st.session_state.auth_name}**"
                   + (" · admin" if st.session_state.auth_is_admin else ""))
    with _sb_col2:
        if st.button("↩", help="Cerrar sesión", use_container_width=True):
            _tok_out = st.session_state.get("auth_token","")
            if _tok_out and _auth_db:
                _auth_db.delete_session(_tok_out)
            st.session_state.authenticated = False
            st.session_state.auth_name     = ""
            st.session_state.auth_username = ""
            st.session_state.auth_is_admin = False
            st.session_state.auth_token    = ""
            st.query_params.clear()
            st.rerun()

    # ── Gestión de usuarios (solo admins) ─────────────────────────────────────
    if st.session_state.auth_is_admin:
        with st.expander("👥 Gestionar usuarios"):
            _mgr_db = get_db()
            if _mgr_db:
                # Lista de usuarios
                try:
                    _all_users = _mgr_db.list_users()
                    for _u in _all_users:
                        _u_col1, _u_col2 = st.columns([4, 1])
                        with _u_col1:
                            st.caption(f"**{_u['name']}** ({_u['username']})"
                                       + (" 🔑" if _u.get("is_admin") else ""))
                        with _u_col2:
                            if _u["username"] != st.session_state.auth_username:
                                if st.button("🗑️", key=f"del_usr_{_u['username']}", help=f"Eliminar {_u['username']}"):
                                    try:
                                        _mgr_db.delete_user(_u["username"])
                                        st.rerun()
                                    except Exception as _eu:
                                        st.error(f"Error: {_eu}")
                except Exception:
                    pass

                st.markdown("**➕ Nuevo usuario**")
                with st.form("new_user_form"):
                    _nu_name  = st.text_input("Nombre", placeholder="Ej: María González", key="nu_name")
                    _nu_user  = st.text_input("Usuario", placeholder="Ej: maria", key="nu_user")
                    _nu_pwd   = st.text_input("Contraseña", type="password", key="nu_pwd")
                    _nu_admin = st.checkbox("Es administrador", key="nu_admin")
                    _nu_ok    = st.form_submit_button("Crear usuario", type="primary", use_container_width=True)
                if _nu_ok:
                    if not _nu_name or not _nu_user or not _nu_pwd:
                        st.error("Todos los campos son obligatorios.")
                    elif len(_nu_pwd) < 6:
                        st.error("Contraseña mínimo 6 caracteres.")
                    else:
                        try:
                            _mgr_db.create_user(_nu_user.strip().lower(), _nu_name.strip(), _hash_pw(_nu_pwd), _nu_admin)
                            st.success(f"✅ Usuario **{_nu_user}** creado.")
                            st.rerun()
                        except Exception as _eu2:
                            st.error(f"Error: {_eu2}")
            else:
                st.warning("Supabase no configurado.")

    st.divider()

    DEMO = st.toggle("⚠️ Modo Demo", value=True,
                     help="Usa datos ficticios. No requiere API keys.")

    st.divider()

    # ── Selector de cliente ───────────────────────────────────────────────────
    st.markdown("**Cliente activo**")
    clients = st.session_state.clients_list or []

    if clients:
        options  = {c["id"]: c["name"] for c in clients}
        names    = ["— Seleccionar —"] + list(options.values())
        ids      = [None] + list(options.keys())
        current  = st.session_state.selected_client_id
        idx      = ids.index(current) if current in ids else 0

        sel_idx = st.selectbox("", names, index=idx, label_visibility="collapsed",
                               key="client_selector")
        new_id  = ids[names.index(sel_idx)]

        if new_id and new_id != st.session_state.selected_client_id:
            client_obj = next((c for c in clients if c["id"] == new_id), None)
            if client_obj:
                select_client(client_obj)
                st.query_params["c"] = str(new_id)
                st.rerun()
    else:
        st.caption("No hay clientes aún.")

    if st.button("＋ Nuevo cliente", use_container_width=True):
        st.session_state.show_client_form  = True
        st.session_state.editing_client_id = None

    if st.session_state.get("auth_is_admin"):
        st.divider()
        if st.session_state.get("view_mode") == "dashboard":
            if st.button("← Volver al Pipeline", use_container_width=True, key="btn_back_pipeline"):
                st.session_state.view_mode = "pipeline"
                st.rerun()
        else:
            if st.button("📊 Dashboard Admin", use_container_width=True, key="btn_go_dashboard"):
                st.session_state.view_mode = "dashboard"
                st.rerun()

    st.divider()

    # ── Estado APIs ───────────────────────────────────────────────────────────
    st.markdown("**Estado de APIs (SOi Digital)**")
    for name, key in [("Claude AI", ANTHROPIC_API_KEY),
                      ("Lusha",     LUSHA_API_KEY),
                      ("Evaboot",   EVABOOT_API_KEY),
                      ("Supabase",  SUPABASE_KEY)]:
        if key:         icon, lbl = "✅", name
        elif DEMO:      icon, lbl = "⚠️", f"{name} (demo)"
        else:           icon, lbl = "❌", name
        st.markdown(f"{icon} &nbsp; {lbl}", unsafe_allow_html=True)

    if st.session_state.selected_client:
        st.markdown("")
        st.markdown("**Lemlist del cliente**")
        lk = lemlist_key()
        if lk:  st.markdown("✅ &nbsp; Lemlist", unsafe_allow_html=True)
        else:   st.markdown("❌ &nbsp; Lemlist (falta key)", unsafe_allow_html=True)

    if DEMO: st.caption("En Modo Demo las APIs no se usan.")

    st.divider()

    # ── Progreso ──────────────────────────────────────────────────────────────
    st.markdown("**Progreso del pipeline**")
    for name, done in [("ICP",            st.session_state.done_icp),
                       ("Buyer Persona",   st.session_state.done_bp),
                       ("Empresas",        st.session_state.done_empresas),
                       ("Contactos",       st.session_state.done_clay),
                       ("Enriquecimiento", st.session_state.done_enrich)]:
        st.markdown(f"{'✅' if done else '⬜'} {name}")

    st.divider()
    st.caption("v2.0 · SOi Digital")

# ══════════════════════════════════════════════════════════════════════════════
# FORMULARIO DE CLIENTE (modal overlay)
# ══════════════════════════════════════════════════════════════════════════════
if st.session_state.show_client_form:
    editing = st.session_state.editing_client_id
    clients = st.session_state.clients_list or []
    editing_data = next((c for c in clients if c["id"] == editing), {}) if editing else {}

    with st.container(border=True):
        title = f"✏️ Editar cliente: {editing_data.get('name','')}" if editing else "➕ Nuevo cliente"
        st.subheader(title)

        col1, col2 = st.columns(2)
        with col1:
            cname = st.text_input("Nombre del cliente *",
                                  value=editing_data.get("name",""),
                                  placeholder="Ej: Acme Corp")
            lemlist_k = st.text_input("API Key de Lemlist",
                                      value=editing_data.get("lemlist_api_key",""),
                                      type="password",
                                      help="Cada cliente tiene su propia cuenta de Lemlist")
        with col2:
            particularidades_val = st.text_area(
                "📝 Particularidades de prospección",
                value=editing_data.get("particularidades_prospeccion",""),
                placeholder='Ej: "Agregar columna con el Área de trabajo del contacto en base a su cargo" · "Agregar columna Región del país del contacto"',
                help='Indícame qué particularidades tiene este cliente para el Excel de exportación. La IA las aplicará automáticamente. Ej: "Crear columna con el Área del contacto según su cargo", "Agregar columna Región del país".',
                height=120,
            )
            st.caption("💡 Las APIs de Evaboot y Lusha son globales (configuradas en Secrets).")

        # ── Empresas a excluir (solo en modo edición) ─────────────────────────
        if editing:
            st.markdown("---")
            st.markdown("##### 🚫 Empresas a excluir de recomendaciones IA")
            st.caption("Sube un Excel con empresas que ya estás trabajando. La IA **nunca** las volverá a recomendar. Columnas necesarias: **nombre_empresa** y/o **dominio_web**.")
            _excl_current = editing_data.get("exclusion_companies") or []
            if _excl_current:
                st.caption(f"📋 Lista actual: **{len(_excl_current)} empresas** excluidas permanentemente.")
                with st.expander("Ver lista de exclusión"):
                    import pandas as _pd_excl
                    st.dataframe(_pd_excl.DataFrame(_excl_current)[
                        [c for c in ["nombre_empresa","dominio_web"] if c in _pd_excl.DataFrame(_excl_current).columns]
                    ], use_container_width=True, hide_index=True)
            _excl_file = st.file_uploader(
                "Subir / reemplazar Excel de exclusión",
                type=["xlsx","xls","csv"],
                key="excl_companies_file",
                help="El Excel anterior se reemplazará. Columnas: nombre_empresa, dominio_web.",
            )
            if _excl_file:
                try:
                    import pandas as _pd_excl2
                    _df_excl = _pd_excl2.read_excel(_excl_file) if not _excl_file.name.endswith(".csv") else _pd_excl2.read_csv(_excl_file)
                    _df_excl.columns = [c.strip().lower().replace(" ","_") for c in _df_excl.columns]
                    _col_alias = {
                        "empresa":"nombre_empresa","company":"nombre_empresa","name":"nombre_empresa",
                        "nombre_de_la_empresa":"nombre_empresa","company_name":"nombre_empresa",
                        "account_name":"nombre_empresa","cuenta":"nombre_empresa",
                        "dominio":"dominio_web","domain":"dominio_web","website":"dominio_web",
                        "url":"dominio_web","url_del_sitio_web":"dominio_web",
                        "sitio_web":"dominio_web","web":"dominio_web","homepage":"dominio_web",
                    }
                    _df_excl.rename(columns={k:v for k,v in _col_alias.items() if k in _df_excl.columns}, inplace=True)
                    # Si aún no hay nombre_empresa, usar la primera columna de texto como fallback
                    if "nombre_empresa" not in _df_excl.columns:
                        _first_txt = next((c for c in _df_excl.columns if _df_excl[c].dtype == object), None)
                        if _first_txt:
                            _df_excl.rename(columns={_first_txt: "nombre_empresa"}, inplace=True)
                    _excl_parsed = []
                    for _, _row in _df_excl.fillna("").iterrows():
                        _e = {
                            "nombre_empresa": str(_row.get("nombre_empresa","")).strip(),
                            "dominio_web"   : str(_row.get("dominio_web","")).strip().lower(),
                        }
                        if _e["nombre_empresa"] or _e["dominio_web"]:
                            _excl_parsed.append(_e)
                    if _excl_parsed:
                        # Calcular cuántas son nuevas (sin duplicar por dominio ni por nombre)
                        _excl_existentes = editing_data.get("exclusion_companies") or []
                        _doms_exist = {e.get("dominio_web","").lower() for e in _excl_existentes if e.get("dominio_web")}
                        _noms_exist = {e.get("nombre_empresa","").lower() for e in _excl_existentes if e.get("nombre_empresa")}
                        _nuevas = [e for e in _excl_parsed
                                   if e.get("dominio_web","").lower() not in _doms_exist
                                   and e.get("nombre_empresa","").lower() not in _noms_exist]
                        _total_merged = len(_excl_existentes) + len(_nuevas)
                        if _nuevas:
                            st.success(f"✅ **{len(_nuevas)} empresas nuevas** para agregar · Total tras guardar: **{_total_merged}**")
                        else:
                            st.info("ℹ️ Todas las empresas del archivo ya estaban en la lista de exclusión.")
                        if st.button("💾 Agregar a lista de exclusión", key="save_excl_btn", type="primary"):
                            _db_excl = get_db()
                            if _db_excl:
                                _merged = _excl_existentes + _nuevas
                                _db_excl.update_client(editing, {"exclusion_companies": _merged})
                                # Actualizar en session_state
                                if st.session_state.selected_client:
                                    st.session_state.selected_client["exclusion_companies"] = _merged
                                _load_idx = next((i for i,c in enumerate(st.session_state.clients_list or []) if c.get("id")==editing), None)
                                if _load_idx is not None:
                                    st.session_state.clients_list[_load_idx]["exclusion_companies"] = _merged
                                st.success(f"✅ Lista actualizada: **{len(_merged)} empresas** excluidas en total.")
                            else:
                                st.warning("Supabase no configurado.")
                    else:
                        st.warning("No se encontraron empresas en el archivo.")
                except Exception as _exc_err:
                    st.error(f"Error leyendo el archivo: {_exc_err}")

        # ── Clientes actuales (Lookalike) ───────────────────────────────────────
        if editing:
            st.markdown("---")
            st.markdown("##### 🎯 Clientes actuales (Lookalike)")
            st.caption("Sube un Excel con los clientes actuales de este cliente. La IA usará estas empresas como referencia para recomendar empresas **similares** (modelo lookalike). Columnas sugeridas: **nombre_empresa**, **industria**, **pais**, **tamano_empleados**, **dominio_web**.")
            _look_current = editing_data.get("lookalike_companies") or []
            if _look_current:
                st.caption(f"📋 Lista actual: **{len(_look_current)} empresas** de referencia lookalike.")
                with st.expander("Ver lista lookalike"):
                    import pandas as _pd_look
                    _df_look_show = _pd_look.DataFrame(_look_current)
                    _look_cols = [c for c in ["nombre_empresa","industria","pais","tamano_empleados","dominio_web"] if c in _df_look_show.columns]
                    st.dataframe(_df_look_show[_look_cols], use_container_width=True, hide_index=True)
                if st.button("🗑️ Eliminar lista lookalike", key="del_look_btn"):
                    _db_look_del = get_db()
                    if _db_look_del:
                        _db_look_del.update_client(editing, {"lookalike_companies": []})
                        if st.session_state.selected_client:
                            st.session_state.selected_client["lookalike_companies"] = []
                        _look_del_idx = next((i for i, c in enumerate(st.session_state.clients_list or []) if c.get("id") == editing), None)
                        if _look_del_idx is not None:
                            st.session_state.clients_list[_look_del_idx]["lookalike_companies"] = []
                        st.success("✅ Lista lookalike eliminada.")
                        st.rerun()
            _look_file = st.file_uploader(
                "Subir / reemplazar Excel de clientes actuales",
                type=["xlsx", "xls", "csv"],
                key="look_companies_file",
                help="La lista anterior se reemplazará. Columnas sugeridas: nombre_empresa, dominio_web, industria, pais, tamano_empleados.",
            )
            if _look_file:
                try:
                    import pandas as _pd_look2
                    _df_look = _pd_look2.read_excel(_look_file) if not _look_file.name.endswith(".csv") else _pd_look2.read_csv(_look_file)
                    _df_look.columns = [c.strip().lower().replace(" ", "_") for c in _df_look.columns]
                    _col_alias_look = {
                        "empresa": "nombre_empresa", "company": "nombre_empresa", "name": "nombre_empresa",
                        "nombre_de_la_empresa": "nombre_empresa", "company_name": "nombre_empresa",
                        "account_name": "nombre_empresa", "cuenta": "nombre_empresa", "cliente": "nombre_empresa",
                        "dominio": "dominio_web", "domain": "dominio_web", "website": "dominio_web",
                        "url": "dominio_web", "web": "dominio_web", "sitio_web": "dominio_web",
                        "industry": "industria", "sector": "industria",
                        "country": "pais", "country_code": "pais", "location": "pais", "ubicacion": "pais",
                        "employees": "tamano_empleados", "num_employees": "tamano_empleados",
                        "headcount": "tamano_empleados", "employee_count": "tamano_empleados", "size": "tamano_empleados",
                    }
                    _df_look.rename(columns={k: v for k, v in _col_alias_look.items() if k in _df_look.columns}, inplace=True)
                    if "nombre_empresa" not in _df_look.columns:
                        _first_txt_look = next((c for c in _df_look.columns if _df_look[c].dtype == object), None)
                        if _first_txt_look:
                            _df_look.rename(columns={_first_txt_look: "nombre_empresa"}, inplace=True)
                    _look_parsed = []
                    for _, _row in _df_look.fillna("").iterrows():
                        _e_look = {
                            "nombre_empresa"  : str(_row.get("nombre_empresa", "")).strip(),
                            "dominio_web"     : str(_row.get("dominio_web", "")).strip().lower(),
                            "industria"       : str(_row.get("industria", "")).strip(),
                            "pais"            : str(_row.get("pais", "")).strip(),
                            "tamano_empleados": str(_row.get("tamano_empleados", "")).strip(),
                        }
                        _e_look = {k: v for k, v in _e_look.items() if v and v != "nan"}
                        if _e_look.get("nombre_empresa") or _e_look.get("dominio_web"):
                            _look_parsed.append(_e_look)
                    if _look_parsed:
                        st.success(f"✅ **{len(_look_parsed)} empresas** de referencia encontradas en el archivo.")
                        if st.button("💾 Guardar lista lookalike", key="save_look_btn", type="primary"):
                            _db_look = get_db()
                            if _db_look:
                                _db_look.update_client(editing, {"lookalike_companies": _look_parsed})
                                if st.session_state.selected_client:
                                    st.session_state.selected_client["lookalike_companies"] = _look_parsed
                                _look_save_idx = next((i for i, c in enumerate(st.session_state.clients_list or []) if c.get("id") == editing), None)
                                if _look_save_idx is not None:
                                    st.session_state.clients_list[_look_save_idx]["lookalike_companies"] = _look_parsed
                                st.success(f"✅ Lista lookalike guardada: **{len(_look_parsed)} empresas** de referencia.")
                            else:
                                st.warning("Supabase no configurado.")
                    else:
                        st.warning("No se encontraron empresas en el archivo.")
                except Exception as _look_err:
                    st.error(f"Error leyendo el archivo: {_look_err}")

        c1, c2, c3 = st.columns([2,1,1])
        with c1:
            if st.button("💾 Guardar cliente", type="primary"):
                if not cname:
                    st.error("El nombre es obligatorio.")
                else:
                    db = get_db()
                    data = {
                        "name"                      : cname,
                        "lemlist_api_key"            : lemlist_k,
                        "particularidades_prospeccion": particularidades_val,
                    }
                    try:
                        if editing and db:
                            updated = db.update_client(editing, data)
                            st.success(f"✅ Cliente '{cname}' actualizado")
                            # Al EDITAR el mismo cliente, solo actualizar metadatos
                            # sin resetear empresas/contactos del pipeline actual
                            _updated_client = {**editing_data, **data,
                                               "id": editing or updated.get("id","")}
                            st.session_state.selected_client = _updated_client
                            # Actualizar solo la key en session state (sin reset de pipeline)
                            load_clients()
                        elif db:
                            updated = db.create_client(data)
                            st.success(f"✅ Cliente '{cname}' creado")
                            select_client({**editing_data, **data,
                                           "id": editing or updated.get("id","")})
                            load_clients()
                        else:
                            updated = {"id":"demo-id","name":cname,**data}
                            select_client({**editing_data, **data,
                                           "id": editing or updated.get("id","")})
                        st.session_state.show_client_form  = False
                        st.session_state.editing_client_id = None
                        st.rerun()
                    except Exception as e:
                        st.error(f"Error guardando: {e}")
        with c2:
            if st.button("Cancelar"):
                st.session_state.show_client_form  = False
                st.session_state.editing_client_id = None
                st.rerun()
        with c3:
            if editing and st.button("🗑️ Eliminar", type="secondary"):
                db = get_db()
                try:
                    if db: db.delete_client(editing)
                    load_clients()
                    st.session_state.selected_client_id = None
                    st.session_state.selected_client    = None
                    st.session_state.show_client_form   = False
                    st.rerun()
                except Exception as e:
                    st.error(f"Error eliminando: {e}")

    st.divider()

# ══════════════════════════════════════════════════════════════════════════════
# HEADER
# ══════════════════════════════════════════════════════════════════════════════
client_name = st.session_state.selected_client["name"] if st.session_state.selected_client else None

st.markdown("# 🎯 Prospector App by BullsEye")
if client_name:
    col_title, col_edit = st.columns([6,1])
    with col_title:
        st.caption(f"Cliente activo: **{client_name}**  ·  Clay · Lusha · Lemlist · Claude AI")
    with col_edit:
        if st.button("✏️ Editar cliente"):
            st.session_state.show_client_form  = True
            st.session_state.editing_client_id = st.session_state.selected_client_id
            st.rerun()
else:
    st.caption("Selecciona o crea un cliente en el sidebar para comenzar")

# Métricas rápidas
import pandas as _pd
m1,m2,m3,m4 = st.columns(4)
with m1:
    st.markdown(f"""<div class="metric-box"><div class="val">{len(st.session_state.empresas)}</div>
    <div class="lbl">Empresas</div></div>""", unsafe_allow_html=True)
with m2:
    st.markdown(f"""<div class="metric-box"><div class="val">{len(st.session_state.contactos_clay)}</div>
    <div class="lbl">Contactos</div></div>""", unsafe_allow_html=True)
with m3:
    emails=sum(1 for c in st.session_state.contactos_final if c.get("email"))
    st.markdown(f"""<div class="metric-box"><div class="val">{emails}</div>
    <div class="lbl">Con Email</div></div>""", unsafe_allow_html=True)
with m4:
    phones=sum(1 for c in st.session_state.contactos_final if c.get("phone"))
    st.markdown(f"""<div class="metric-box"><div class="val">{phones}</div>
    <div class="lbl">Con Teléfono</div></div>""", unsafe_allow_html=True)

st.divider()

# ══════════════════════════════════════════════════════════════════════════════
# VISTA DASHBOARD ADMIN
# ══════════════════════════════════════════════════════════════════════════════
if st.session_state.get("view_mode") == "dashboard":
    import pandas as _pd_dash
    st.markdown("## 📊 Dashboard Administrativo")
    st.caption("Vista general de toda la actividad en la plataforma.")

    _dash_db = get_db()
    if not _dash_db:
        st.warning("Supabase no configurado.")
    else:
        try:
            _all_runs    = (_dash_db.get_all_runs() if hasattr(_dash_db, "get_all_runs") else []) or []
            _all_clients = _dash_db.list_clients() or []
            _usage_logs  = (_dash_db.get_usage_logs() if hasattr(_dash_db, "get_usage_logs") else []) or []
            _client_map  = {c["id"]: c.get("name","?") for c in _all_clients}

            # Filtro por cliente
            _filt_opts = ["Todos los clientes"] + [c.get("name","?") for c in _all_clients]
            _filt_sel  = st.selectbox("🔍 Filtrar por cliente:", _filt_opts, key="dash_client_filter")
            if _filt_sel != "Todos los clientes":
                _filt_id = next((c["id"] for c in _all_clients if c.get("name") == _filt_sel), None)
                _all_runs    = [r for r in _all_runs if r.get("client_id") == _filt_id]
                _usage_logs  = [u for u in _usage_logs if u.get("client_name") == _filt_sel]

            st.divider()

            # ── KPIs generales ────────────────────────────────────────────────
            st.markdown("### 🔢 Resumen general")
            _total_clients  = len(_all_clients)
            _total_runs     = len(_all_runs)
            _total_contacts = sum((r.get("stats") or {}).get("total",0) for r in _all_runs)
            _total_emails   = sum((r.get("stats") or {}).get("emails",0) for r in _all_runs)
            _total_phones   = sum((r.get("stats") or {}).get("phones",0) for r in _all_runs)
            _total_lusha    = sum((r.get("stats") or {}).get("lusha",0) for r in _all_runs)

            _k1,_k2,_k3,_k4,_k5,_k6 = st.columns(6)
            _k1.metric("Clientes activos",      _total_clients)
            _k2.metric("Prospecciones totales", _total_runs)
            _k3.metric("Contactos totales",     _total_contacts)
            _k4.metric("Emails encontrados",    _total_emails,
                       f"{_total_emails/_total_contacts*100:.0f}%" if _total_contacts else "")
            _k5.metric("Teléfonos totales",     _total_phones,
                       f"{_total_phones/_total_contacts*100:.0f}%" if _total_contacts else "")
            _k6.metric("Via Lusha",             _total_lusha)

            st.divider()

            # ── Actividad por prospector ──────────────────────────────────────
            st.markdown("### 👤 Actividad por prospector")
            _user_stats = {}
            for _r in _all_runs:
                _uname = (_r.get("stats") or {}).get("user","") or "Sin asignar"
                if _uname not in _user_stats:
                    _user_stats[_uname] = {"runs":0,"contacts":0,"emails":0,"phones":0,"lusha":0}
                _s2 = _r.get("stats") or {}
                _user_stats[_uname]["runs"]     += 1
                _user_stats[_uname]["contacts"] += _s2.get("total",0)
                _user_stats[_uname]["emails"]   += _s2.get("emails",0)
                _user_stats[_uname]["phones"]   += _s2.get("phones",0)
                _user_stats[_uname]["lusha"]    += _s2.get("lusha",0)
            if _user_stats:
                st.dataframe(_pd_dash.DataFrame([
                    {"Prospector": u, "Prospecciones": v["runs"],
                     "Contactos": v["contacts"], "Emails": v["emails"],
                     "Teléfonos": v["phones"], "Via Lusha": v["lusha"]}
                    for u, v in sorted(_user_stats.items(), key=lambda x: -x[1]["runs"])
                ]), use_container_width=True, hide_index=True)
            else:
                st.info("Aún no hay prospecciones guardadas con usuario asignado.")

            st.divider()

            # ── Resumen por cliente ───────────────────────────────────────────
            st.markdown("### 🏢 Resumen por cliente")
            _client_stats = {}
            for _r in _all_runs:
                _cid2  = _r.get("client_id","")
                _cnom = _client_map.get(_cid2, _cid2 or "?")
                if _cnom not in _client_stats:
                    _client_stats[_cnom] = {"runs":0,"contacts":0,"emails":0,"phones":0,"lusha":0,"rechazadas":0}
                _s3 = _r.get("stats") or {}
                _client_stats[_cnom]["runs"]     += 1
                _client_stats[_cnom]["contacts"] += _s3.get("total",0)
                _client_stats[_cnom]["emails"]   += _s3.get("emails",0)
                _client_stats[_cnom]["phones"]   += _s3.get("phones",0)
                _client_stats[_cnom]["lusha"]    += _s3.get("lusha",0)
            for _c2 in _all_clients:
                _cnom2 = _c2.get("name","?")
                if _cnom2 not in _client_stats:
                    _client_stats[_cnom2] = {"runs":0,"contacts":0,"emails":0,"phones":0,"lusha":0,"rechazadas":0}
                _client_stats[_cnom2]["rechazadas"] = len(_parse_json_field(_c2.get("empresas_rechazadas")) or [])
            if _client_stats:
                st.dataframe(_pd_dash.DataFrame([
                    {"Cliente": n, "Prospecciones": v["runs"],
                     "Contactos": v["contacts"], "Emails": v["emails"],
                     "Teléfonos": v["phones"], "Via Lusha": v["lusha"],
                     "Emp. Rechazadas": v["rechazadas"]}
                    for n, v in sorted(_client_stats.items(), key=lambda x: -x[1]["contacts"])
                ]), use_container_width=True, hide_index=True)

            st.divider()

            # ── Empresas rechazadas ───────────────────────────────────────────
            st.markdown("### 🚫 Empresas rechazadas (todos los clientes)")
            _rech_rows = []
            for _c3 in _all_clients:
                for _re in (_parse_json_field(_c3.get("empresas_rechazadas")) or []):
                    _rech_rows.append({
                        "Cliente": _c3.get("name","?"),
                        "Empresa": _re.get("nombre_empresa",""),
                        "Dominio": _re.get("dominio_web",""),
                        "Razón":   _re.get("razon_rechazo",""),
                        "Fecha":   _re.get("fecha_rechazo",""),
                    })
            if _rech_rows:
                st.caption(f"**{len(_rech_rows)} empresas rechazadas** en total.")
                st.dataframe(_pd_dash.DataFrame(_rech_rows), use_container_width=True, hide_index=True)
            else:
                st.info("No hay empresas rechazadas registradas.")

            st.divider()

            # ── Uso de servicios ──────────────────────────────────────────────
            st.markdown("### 💳 Uso de servicios")
            if _usage_logs:
                _svc_stats = {}
                for _ul in _usage_logs:
                    _key = f"{_ul.get('service','?')} · {_ul.get('action','?')}"
                    if _key not in _svc_stats:
                        _svc_stats[_key] = {"llamadas":0,"unidades":0}
                    _svc_stats[_key]["llamadas"] += 1
                    _svc_stats[_key]["unidades"] += _ul.get("units",1)
                st.dataframe(_pd_dash.DataFrame([
                    {"Servicio / Acción": k, "Llamadas": v["llamadas"], "Unidades": v["unidades"]}
                    for k, v in sorted(_svc_stats.items(), key=lambda x: -x[1]["llamadas"])
                ]), use_container_width=True, hide_index=True)
                st.markdown("**Detalle reciente (últimas 50 acciones)**")
                st.dataframe(_pd_dash.DataFrame([{
                    "Fecha"   : (_ul.get("created_at","") or "")[:16].replace("T"," "),
                    "Usuario" : _ul.get("user_username",""),
                    "Cliente" : _ul.get("client_name",""),
                    "Servicio": _ul.get("service",""),
                    "Acción"  : _ul.get("action",""),
                    "Unidades": _ul.get("units",1),
                } for _ul in _usage_logs[:50]]), use_container_width=True, hide_index=True)
            else:
                st.info("Aún no hay registros de uso. El log se irá llenando a medida que uses la app.")

            st.divider()

            # ── Actividad reciente ────────────────────────────────────────────
            st.markdown("### 📅 Actividad reciente (últimas prospecciones)")
            if _all_runs:
                st.dataframe(_pd_dash.DataFrame([{
                    "Fecha"     : ((_r.get("run_date") or (_r.get("stats") or {}).get("date","") or "")[:10]),
                    "Cliente"   : _client_map.get(_r.get("client_id",""), "?"),
                    "Prospector": (_r.get("stats") or {}).get("user","") or "—",
                    "Contactos" : (_r.get("stats") or {}).get("total",0),
                    "Emails"    : (_r.get("stats") or {}).get("emails",0),
                    "Teléfonos" : (_r.get("stats") or {}).get("phones",0),
                    "Via Lusha" : (_r.get("stats") or {}).get("lusha",0),
                } for _r in _all_runs[:30]]), use_container_width=True, hide_index=True)
            else:
                st.info("No hay prospecciones guardadas aún.")

        except Exception as _dash_err:
            st.error(f"Error cargando el dashboard: {_dash_err}")
            import traceback
            st.code(traceback.format_exc())
    st.stop()

# ══════════════════════════════════════════════════════════════════════════════
# PANTALLA DE BIENVENIDA (sin cliente seleccionado)
# ══════════════════════════════════════════════════════════════════════════════
if not st.session_state.selected_client and not st.session_state.show_client_form:
    st.markdown("### 👋 Bienvenido a Prospector App by BullsEye")
    clients = st.session_state.clients_list or []
    if clients:
        st.markdown("Selecciona un cliente en el sidebar para comenzar el pipeline, "
                    "o crea uno nuevo.")
        st.markdown("#### Clientes disponibles")
        for c in clients:
            col_n, col_btn = st.columns([4,1])
            with col_n:
                icp_ok = "✅" if c.get("icp") else "⬜"
                bp_ok  = "✅" if c.get("buyer_persona") else "⬜"
                lem_ok = "✅" if c.get("lemlist_api_key") else "❌"
                st.markdown(f"**{c['name']}** &nbsp; ICP {icp_ok} &nbsp; BP {bp_ok} "
                            f"&nbsp; Lemlist {lem_ok}", unsafe_allow_html=True)
            with col_btn:
                if st.button("Seleccionar", key=f"sel_{c['id']}"):
                    select_client(c); st.rerun()
    else:
        st.info("No hay clientes aún. Crea el primero usando el botón **＋ Nuevo cliente** "
                "en el sidebar.")
    st.stop()

# ══════════════════════════════════════════════════════════════════════════════
# TABS DEL PIPELINE
# ══════════════════════════════════════════════════════════════════════════════
tab0,tab1,tab2,tab3,tab4,tab5,tab6,tab7,tab8 = st.tabs([
    "💡  Propuesta de Valor",
    "🎯  ICP",
    "👤  Buyer Persona",
    "🏢  Empresas",
    "👥  Contactos",
    "📱  Enriquecimiento",
    "📊  Resultados",
    "🕐  Historial",
    "🚫  Rechazadas",
])

# ── TAB 0 · PROPUESTA DE VALOR ────────────────────────────────────────────────
with tab0:
    st.subheader("Propuesta de Valor")
    saved_pv = st.session_state.propuesta_de_valor or {}

    st.markdown("##### Sitio web de la empresa")
    url_col, btn_col = st.columns([3, 1])
    with url_col:
        empresa_url = st.text_input(
            "URL del sitio web",
            value=saved_pv.get("url", ""),
            placeholder="https://www.tuempresa.com",
            label_visibility="collapsed",
        )
    with btn_col:
        analizar_btn = st.button("🤖 Analizar con IA", key="btn_analizar_web",
                                 use_container_width=True)

    if analizar_btn:
        if not empresa_url.strip():
            st.warning("Ingresa la URL del sitio web primero.")
        elif not ANTHROPIC_API_KEY:
            st.error("No hay ANTHROPIC_API_KEY configurada.")
        else:
            with st.spinner("Leyendo el sitio web y generando propuesta de valor…"):
                try:
                    resultado_ia = analizar_web_empresa(empresa_url.strip())
                    # Guardar en session state para pre-llenar los campos
                    st.session_state.propuesta_de_valor = {
                        "url"       : empresa_url.strip(),
                        "propuesta" : resultado_ia.get("propuesta_de_valor", ""),
                        "dolores"   : "\n".join(resultado_ia.get("dolores_que_soluciona", [])),
                        # Guardamos sugerencias de BP para usarlas en Tab 2
                        "_cargos_sugeridos"      : resultado_ia.get("cargos_sugeridos", []),
                        "_roles_compra"          : resultado_ia.get("roles_compra", {}),
                        "_pain_points_sugeridos" : resultado_ia.get("pain_points_buyer", []),
                    }
                    st.success("✅ Análisis completado. Revisa y edita los campos abajo.")
                    st.rerun()
                except Exception as e:
                    _err = str(e)
                    if "529" in _err or "overloaded" in _err.lower():
                        st.warning("⚠️ La API de Anthropic está sobrecargada en este momento. "
                                   "El sistema reintentó automáticamente (5 intentos + modelo alternativo). "
                                   "**Espera 1-2 minutos y vuelve a hacer clic en 'Analizar con IA'.**")
                    else:
                        st.error(f"Error al analizar el sitio: {e}")

    st.markdown("---")
    pv_propuesta = st.text_area(
        "Propuesta de valor",
        value=saved_pv.get("propuesta", ""),
        max_chars=300,
        height=100,
        placeholder="Describe qué hace tu empresa y qué valor único entrega a sus clientes (máx. 300 caracteres).",
        help="Máximo 300 caracteres.",
    )
    chars_left = 300 - len(pv_propuesta)
    st.caption(f"{chars_left} caracteres restantes")

    pv_dolores = st.text_area(
        "Dolores que soluciona",
        value=saved_pv.get("dolores", ""),
        height=150,
        placeholder="Lista los principales problemas o dolores que tu solución resuelve (uno por línea).",
    )

    if st.button("💾 Guardar Propuesta de Valor", type="primary", key="save_pv"):
        if not pv_propuesta.strip():
            st.error("Completa la propuesta de valor.")
        else:
            new_pv = {
                "url"      : empresa_url.strip(),
                "propuesta": pv_propuesta.strip(),
                "dolores"  : pv_dolores.strip(),
                # Preservar sugerencias de IA si existen
                "_cargos_sugeridos"      : saved_pv.get("_cargos_sugeridos", []),
                "_roles_compra"          : saved_pv.get("_roles_compra", {}),
                "_pain_points_sugeridos" : saved_pv.get("_pain_points_sugeridos", []),
            }
            st.session_state.propuesta_de_valor = new_pv
            st.session_state.done_pv = True
            db = get_db()
            if db and st.session_state.selected_client_id:
                try:
                    db.update_client(st.session_state.selected_client_id,
                                     {"propuesta_de_valor": new_pv})
                    st.session_state.selected_client["propuesta_de_valor"] = new_pv
                    st.success(f"✅ Propuesta de valor guardada para **{client_name}**. Continúa en ICP →")
                except Exception as _epv:
                    st.error(f"❌ Error al guardar en Supabase: {_epv}")
            else:
                st.success(f"✅ Propuesta de valor guardada para **{client_name}**. Continúa en ICP →")


# ── TAB 1 · ICP ───────────────────────────────────────────────────────────────
with tab1:
    if True:  # ICP siempre accesible — no requiere Propuesta de Valor completada
        st.subheader("ICP — Perfil de empresa objetivo")
        saved_icp = st.session_state.icp or {}

        # ── IA: Generador de ICP desde texto libre ──────────────────────────────
        with st.expander("🤖 Describe tu empresa ideal y la IA completará el ICP", expanded=False):
            desc_empresa = st.text_area(
                "Describe en tus propias palabras qué tipo de empresa quieres prospectar",
                placeholder=(
                    "Ej: Empresas de tecnología B2B en México y Colombia, con entre 50 y 300 empleados, "
                    "que tengan equipo comercial, usen CRM y hayan levantado capital. "
                    "Que no sean empresas de consumo masivo ni startups sin facturación."
                ),
                height=110,
                key="desc_empresa_ia",
            )
            if st.button("✨ Generar ICP con IA", key="btn_generar_icp"):
                if not desc_empresa.strip():
                    st.warning("Escribe una descripción primero.")
                elif not ANTHROPIC_API_KEY:
                    st.error("No hay ANTHROPIC_API_KEY configurada.")
                else:
                    with st.spinner("Analizando descripción y generando ICP…"):
                        try:
                            icp_ia = generar_icp_con_ia(desc_empresa.strip())
                            st.session_state.icp = icp_ia
                            st.success("✅ ICP generado. Revisa los campos abajo y guarda cuando estés listo.")
                            st.rerun()
                        except Exception as e:
                            _err = str(e)
                            if "529" in _err or "overloaded" in _err.lower():
                                st.warning("⚠️ La IA está temporalmente sobrecargada. Espera unos segundos e intenta de nuevo.")
                            else:
                                st.error(f"Error generando ICP: {e}")

        col_a, col_b = st.columns(2)
        with col_a:
            _ind_opciones = [
                "SaaS B2B","Fintech","Tecnología / Software","Servicios profesionales",
                "E-commerce B2B","Edtech","Healthtech","Logística","HR Tech",
                "Ciberseguridad","Inteligencia Artificial / ML","Marketing Tech",
                "Legal Tech","Proptech / Real Estate","Insurtech","Agritech",
                "Retail / Comercio","Manufactura","Construcción","Energía / Utilities",
                "Telecomunicaciones","Medios / Entretenimiento","Viajes / Turismo",
                "Alimentos y Bebidas","Automotriz","Consultoría empresarial",
                "Contabilidad / ERP","Publicidad / Agencias","BPO / Outsourcing",
                "Salud / Clínicas","Educación corporativa","Gobierno / Sector público",
            ]
            _ind_guardadas = saved_icp.get("industrias", ["SaaS B2B","Fintech"])
            _opciones_final = _ind_opciones + [i for i in _ind_guardadas if i not in _ind_opciones]
            industrias = st.multiselect("Industrias objetivo",
                _opciones_final,
                default=[i for i in _ind_guardadas if i in _opciones_final])
            ind_custom = st.text_input("¿Otra industria? Escríbela aquí",
                placeholder="Ej: Proptech, Cleantech, Retail B2B…",
                help="Se agregará a las industrias seleccionadas al guardar")
            if ind_custom.strip():
                industrias = industrias + [ind_custom.strip()]
            _paises_mundo = [
                "México","Colombia","Argentina","Chile","Perú","Brasil","Uruguay",
                "Ecuador","Costa Rica","Bolivia","Paraguay","Venezuela","Panamá",
                "Guatemala","Honduras","El Salvador","Nicaragua","Cuba","República Dominicana",
                "Puerto Rico","Estados Unidos","Canadá",
                "España","Reino Unido","Alemania","Francia","Italia","Portugal",
                "Países Bajos","Bélgica","Suecia","Noruega","Dinamarca","Finlandia",
                "Suiza","Austria","Polonia","Rumania","República Checa","Hungría",
                "Grecia","Irlanda","Israel",
                "India","China","Japón","Corea del Sur","Singapur","Australia",
                "Nueva Zelanda","Indonesia","Malasia","Vietnam","Tailandia","Filipinas",
                "Emiratos Árabes Unidos","Arabia Saudita","Qatar","Sudáfrica","Egipto","Nigeria",
            ]
            geografias = st.multiselect("Países",
                _paises_mundo,
                default=saved_icp.get("geografias", ["México","Colombia","Argentina","Chile"]))
            tam = saved_icp.get("tamano_empresa",{})
            _EMP_OPCIONES = [10,25,50,100,200,300,500,750,1000,2000,3000,5000,10000,20000,50000]
            _emp_labels   = {v: (f"{v:,}+" if v == 50000 else f"{v:,}") for v in _EMP_OPCIONES}
            _emp_min_def  = tam.get("empleados_min", 50)
            _emp_max_def  = tam.get("empleados_max", 500)
            def _snap(val, opts): return min(opts, key=lambda x: abs(x - val))
            _emp_min_idx = _EMP_OPCIONES.index(_snap(_emp_min_def, _EMP_OPCIONES))
            emp_col1, emp_col2 = st.columns(2)
            with emp_col1:
                _emp_min_sel = st.selectbox("Empleados mínimo",
                    options=_EMP_OPCIONES, index=_emp_min_idx,
                    format_func=lambda v: _emp_labels[v], key="emp_min")
            with emp_col2:
                _emp_max_opts = [v for v in _EMP_OPCIONES if v >= _emp_min_sel]
                _emp_max_idx2 = _emp_max_opts.index(_snap(_emp_max_def, _emp_max_opts)) if _snap(_emp_max_def, _emp_max_opts) in _emp_max_opts else len(_emp_max_opts)-1
                _emp_max_sel = st.selectbox("Empleados máximo",
                    options=_emp_max_opts, index=_emp_max_idx2,
                    format_func=lambda v: f"{v:,}+" if v == 50000 else f"{v:,}",
                    key="emp_max")
            empleados = (_emp_min_sel, _emp_max_sel)
            modelo = st.radio("Modelo de negocio",
                ["B2B","B2B y B2C","Marketplace B2B"], horizontal=True,
                index=["B2B","B2B y B2C","Marketplace B2B"].index(saved_icp.get("modelo_negocio","B2B")))
        with col_b:
            _SENALES_CATALOGO = [
                "Tiene equipo de ventas","Tiene equipo de marketing","Usa CRM",
                "Usa herramientas de sales engagement","Tiene SDRs o BDRs",
                "Está contratando vendedores","Está contratando roles de revenue",
                "Recibió financiamiento reciente","En expansión a nuevos mercados",
                "Abrió nuevas oficinas o sedes","Lanzó un nuevo producto recientemente",
                "Facturación en crecimiento (señales públicas)",
                "Tiene página de precios B2B","Tiene blog o contenido activo",
                "Activo en LinkedIn (publicaciones frecuentes)","Usa publicidad pagada B2B",
                "Tiene caso de éxito / case studies publicados",
                "Asiste a eventos del sector","Es speaker en conferencias",
                "Miembro de asociaciones del sector",
                "Usa Salesforce, HubSpot u otro CRM enterprise",
                "Usa Slack o herramientas de productividad cloud",
                "Tiene API pública o integraciones","Stack tecnológico moderno (SaaS-first)",
            ]

            # Leer default guardado
            _senales_default = saved_icp.get(
                "senales_fit",
                ["Tiene equipo de ventas","Usa CRM","Está contratando vendedores"]
            )

            # ── Pills interactivos: todos los ítems visibles, clic para activar/desactivar ──
            _senales_result = st.pills(
                "Señales de fit — activa o desactiva haciendo clic",
                options=_SENALES_CATALOGO,
                default=_senales_default,
                selection_mode="multi",
                key="icp_senales_pills",
            )
            senales = list(_senales_result) if _senales_result is not None else list(_senales_default)
            excl_default = "\n".join(saved_icp.get("exclusiones",
                ["B2C puro","Startups sin revenue","Empresas sin equipo comercial"]))
            exclusiones = st.text_area("Exclusiones (una por línea)", value=excl_default, height=148)

        # ── Geo ID de LinkedIn (override manual) ─────────────────────────────
        with st.expander("🌐 Geo ID de LinkedIn Sales Navigator (avanzado)", expanded=False):
            st.caption(
                "Si el filtro de país no aplica correctamente en Sales Navigator, "
                "puedes ingresar el ID geográfico correcto aquí.\n\n"
                "**Cómo obtenerlo:** Ve a Sales Navigator → búsqueda de leads → "
                "agrega manualmente el filtro de Geografía con tu país → copia la URL → "
                "busca el número después de `id%3A` en el parámetro `GEOGRAPHY`."
            )
            _geo_id_default = saved_icp.get("linkedin_geo_id", "")
            linkedin_geo_id = st.text_input(
                "Geo ID (dejar vacío para usar el automático)",
                value=_geo_id_default,
                placeholder="Ej: 104621616",
                key="linkedin_geo_id_input",
            )

        if st.button("💾 Guardar ICP", type="primary", key="save_icp"):
            if not industrias or not geografias:
                st.error("Selecciona al menos una industria y un país.")
            else:
                new_icp = {
                    "industrias"    : industrias,
                    "geografias"    : geografias,
                    "tamano_empresa": {"empleados_min":empleados[0],"empleados_max":empleados[1]},
                    "modelo_negocio": modelo,
                    "senales_fit"   : senales,
                    "exclusiones"   : [e.strip() for e in exclusiones.splitlines() if e.strip()],
                    "linkedin_geo_id": linkedin_geo_id.strip() if linkedin_geo_id else "",
                }
                st.session_state.icp      = new_icp
                st.session_state.done_icp = True
                db = get_db()
                if db and st.session_state.selected_client_id:
                    try:
                        db.update_client(st.session_state.selected_client_id, {"icp": new_icp})
                        st.session_state.selected_client["icp"] = new_icp
                    except: pass
                st.success(f"✅ ICP guardado para **{client_name}**. Continúa en Buyer Persona →")

# ── TAB 2 · BUYER PERSONA ─────────────────────────────────────────────────────
with tab2:
    if not st.session_state.done_icp:
        st.info("👈 Primero completa el **ICP**.")
    else:
        st.subheader("Buyer Persona")
        saved_bp = st.session_state.buyer_persona or {}
        # Sugerencias de IA guardadas en Propuesta de Valor
        _pv_data          = st.session_state.propuesta_de_valor or {}
        _cargos_sugeridos = _pv_data.get("_cargos_sugeridos", [])
        _roles_ia         = _pv_data.get("_roles_compra", {})
        _pp_sugeridos     = _pv_data.get("_pain_points_sugeridos", [])

        # ── Lista completa de cargos ─────────────────────────────────────────────
        _todos_cargos = [
            # C-Suite
            "CEO","Co-CEO","Founder","Co-Founder","Owner","Managing Director",
            "President","Executive Director",
            # Revenue / Ventas
            "CRO","Chief Revenue Officer","VP of Sales","VP Comercial","VP Revenue",
            "Director Comercial","Director de Ventas","Director de Revenue",
            "Head of Sales","Head of Revenue","Head of Business Development",
            "Gerente Comercial","Gerente de Ventas","Sales Manager",
            "Regional Sales Manager","Country Manager","Sales Director",
            "Account Executive","Senior Account Executive","Enterprise AE",
            "SDR Manager","Head of SDR","BDR Manager",
            # Marketing
            "CMO","Chief Marketing Officer","VP of Marketing","Director de Marketing",
            "Head of Marketing","Head of Growth","Growth Manager",
            "Demand Generation Manager","Head of Demand Gen",
            # Operaciones / Tecnología
            "COO","Chief Operating Officer","VP of Operations","Director de Operaciones",
            "CTO","Chief Technology Officer","VP of Engineering","Head of IT",
            "CIO","Chief Information Officer","IT Manager","IT Director",
            # Finanzas
            "CFO","Chief Financial Officer","VP of Finance","Finance Director",
            "Finance Manager",
            # Customer Success / Producto
            "VP of Customer Success","Head of Customer Success","CS Manager",
            "CPO","Chief Product Officer","VP of Product","Head of Product",
            # RRHH / Talento
            "CHRO","Chief People Officer","VP of HR","HR Director","HR Manager",
            "Head of Talent","Talent Acquisition Manager",
            # Otros
            "General Manager","Business Development Manager","Partnerships Manager",
            "Head of Strategy","Strategy Director",
        ]
        # Detectar mercado anglófono y agregar títulos en inglés
        _icp_raw    = st.session_state.icp or {}
        _geo_raw    = _icp_raw.get("geografias", [])
        _paises_icp = (_geo_raw if isinstance(_geo_raw, list) else []) or \
                      _icp_raw.get("paises", [])
        _english_markets = {"Estados Unidos","United States","Canadá","Canada",
                            "Reino Unido","United Kingdom","Australia","Irlanda","Nueva Zelanda"}
        _is_english_market = any(p in _english_markets for p in _paises_icp)

        _english_titles = [
            # Industry-specific English titles
            "Lab Owner","Laboratory Director","Lab Manager","Dental Lab Owner",
            "Practice Owner","Clinic Owner","Clinic Director","Medical Director",
            "Hospital Administrator","Healthcare Director","VP of Healthcare",
            "Plant Manager","Operations Director","Supply Chain Director",
            "Manufacturing Director","VP of Manufacturing","Production Manager",
            "Retail Director","Store Manager","Regional Manager",
            "Restaurant Owner","F&B Director","Hospitality Manager",
            "Principal","School Director","Academic Director","Dean",
            "Real Estate Developer","Property Manager","Construction Manager",
            "Marketing Manager","Digital Marketing Manager","Brand Manager",
            "E-commerce Manager","Head of Digital","Digital Director",
            "Procurement Manager","Purchasing Director","Category Manager",
        ] if _is_english_market else []

        # Añadir cargos guardados, sugeridos y títulos en inglés que no estén en la lista
        _cargos_guardados = saved_bp.get("cargos_objetivo", [])
        _extra = [c for c in _cargos_guardados + _cargos_sugeridos + _english_titles
                  if c not in _todos_cargos]
        _todos_cargos_final = _todos_cargos + _extra

        if _is_english_market:
            st.caption("💡 Mercado anglófono detectado — se agregaron títulos en inglés a la lista.")

        # ── Default inicial de cargos ────────────────────────────────────────────
        _cargos_default = _cargos_guardados if _cargos_guardados else \
                          (_cargos_sugeridos if _cargos_sugeridos else
                           ["CEO","Founder","VP of Sales","Director Comercial"])
        _cargos_default_valid = [c for c in _cargos_default if c in _todos_cargos_final] \
                                 or [c for c in _cargos_default]  # incluir aunque sea custom

        # Estado de cargos seleccionados (por sesión; se reinicia al cambiar cliente)
        if "bp_cargos_sel" not in st.session_state:
            st.session_state.bp_cargos_sel = list(_cargos_default_valid)
        _sel = st.session_state.bp_cargos_sel

        # ── Pills interactivos: haz clic en uno para quitarlo ───────────────────
        st.markdown("**Cargos objetivo** — haz clic en un cargo para quitarlo")
        if _sel:
            # Clave dinámica para forzar reset cuando cambia la lista
            _pills_key_c = f"pills_cargos_{hash(tuple(sorted(_sel)))}"
            _pills_result = st.pills(
                "cargos_pills",
                options=list(_sel),
                default=list(_sel),
                selection_mode="multi",
                key=_pills_key_c,
                label_visibility="collapsed",
            )
            # Detectar si el usuario quitó alguno (deseleccionó)
            if _pills_result is not None and set(_pills_result) != set(_sel):
                st.session_state.bp_cargos_sel = list(_pills_result)
                st.rerun()
        else:
            st.caption("Ningún cargo seleccionado — busca abajo para agregar.")

        # ── Buscador para agregar cargos ─────────────────────────────────────────
        cargo_search = st.text_input("🔍 Buscar cargo para agregar",
                                     placeholder="Ej: Sales, CTO, Marketing…",
                                     key="cargo_search")
        if cargo_search:
            _match = [c for c in _todos_cargos_final
                      if cargo_search.lower() in c.lower() and c not in _sel]
            _texto_libre = cargo_search.strip()
            _ya_en_lista = _texto_libre in _sel

            # Sugerencias del catálogo
            if _match:
                _add_cols = st.columns(min(3, len(_match)))
                for _mi, _opt in enumerate(_match[:6]):
                    if _add_cols[_mi % min(3, len(_match))].button(
                            f"＋ {_opt}", key=f"add_c_{_opt}", use_container_width=True):
                        _sel.append(_opt)
                        st.rerun()

            # Opción de texto libre — siempre visible si no está ya en la lista
            if not _ya_en_lista:
                if st.button(
                    f'＋ Agregar "{_texto_libre}" exactamente como está',
                    key="add_custom_cargo",
                    help="Agrega el texto tal como lo escribiste, sin modificarlo",
                ):
                    _sel.append(_texto_libre)
                    st.rerun()
            else:
                st.caption(f"✅ «{_texto_libre}» ya está en la lista.")

        cargos = _sel  # compatibilidad con el resto del código

        if _cargos_sugeridos and not saved_bp.get("cargos_objetivo"):
            st.caption(f"✨ La IA sugirió: {', '.join(_cargos_sugeridos[:5])}")

        # ── Palabras clave / cargos a EXCLUIR ────────────────────────────────────
        st.markdown("**🚫 Palabras clave a excluir de la búsqueda**")
        _excluidos_guardados = "\n".join(saved_bp.get("cargos_excluidos", []))
        cargos_excluir_raw = st.text_area(
            "Cargos o palabras clave a excluir (una por línea)",
            value=_excluidos_guardados,
            height=90,
            placeholder="ejecutivo\njunior\nasistente\npracticante",
            help=(
                "Estas palabras se agregarán como NOT en Sales Navigator.\n"
                "Ejemplo: si pones 'junior' y 'ejecutivo', la búsqueda será\n"
                "(VP de Ventas OR Director) NOT (junior OR ejecutivo)"
            ),
            key="cargos_excluir_input",
        )
        _cargos_excluidos = [k.strip() for k in cargos_excluir_raw.splitlines() if k.strip()]
        if _cargos_excluidos:
            st.caption(f"🚫 Se excluirá: {' · '.join(_cargos_excluidos)}")

        # ── Rol de compra ────────────────────────────────────────────────────────
        _roles_default = saved_bp.get("roles_compra", [])
        if not _roles_default:
            _roles_default = []
            if _roles_ia.get("tomadores_de_decision"): _roles_default.append("Tomadores de decisión")
            if _roles_ia.get("influenciadores"):        _roles_default.append("Influenciadores")
        roles_compra = st.multiselect(
            "Rol de compra",
            ["Tomadores de decisión", "Influenciadores"],
            default=_roles_default if _roles_default else ["Tomadores de decisión"],
            help="Define si te diriges a quienes aprueban la compra, a quienes influyen en ella, o a ambos.",
        )
        if _roles_ia:
            with st.expander("Ver detalle de roles sugeridos por IA", expanded=False):
                if _roles_ia.get("tomadores_de_decision"):
                    st.markdown(f"**Tomadores de decisión:** {', '.join(_roles_ia['tomadores_de_decision'])}")
                if _roles_ia.get("influenciadores"):
                    st.markdown(f"**Influenciadores:** {', '.join(_roles_ia['influenciadores'])}")

        st.markdown("---")

        # ── Pain points ──────────────────────────────────────────────────────────
        _pp_guardados = saved_bp.get("pain_points", [])
        _pp_usar = _pp_guardados if _pp_guardados else _pp_sugeridos
        pp_default = "\n".join(_pp_usar) if _pp_usar else (
            "Prospección manual consume demasiado tiempo\n"
            "Pipeline insuficiente o poco calificado\n"
            "Bajo ratio de respuesta en outreach"
        )
        if _pp_sugeridos and not _pp_guardados:
            st.caption("✨ Pain points pre-completados por la IA en base a la propuesta de valor.")
        pain_points = st.text_area("Pain points (uno por línea)", value=pp_default, height=140)

        # ── Empresas a recomendar ─────────────────────────────────────────────────
        n_empresas = st.slider("Empresas a recomendar", 5, 1000,
                               st.session_state.n_empresas, step=5)

        if st.button("💾 Guardar Buyer Persona", type="primary", key="save_bp"):
            if not cargos:
                st.error("Selecciona al menos un cargo.")
            else:
                new_bp = {
                    "cargos_objetivo" : cargos,
                    "cargos_excluidos": _cargos_excluidos,
                    "roles_compra"    : roles_compra,
                    "pain_points"     : [p.strip() for p in pain_points.splitlines() if p.strip()],
                }
                new_cr = {
                    "tiene_equipo_ventas": True,
                    "empleados_minimo"   : st.session_state.icp["tamano_empresa"]["empleados_min"],
                    "usa_crm"            : True,
                }
                st.session_state.buyer_persona = new_bp
                st.session_state.criterios     = new_cr
                st.session_state.n_empresas    = n_empresas
                st.session_state.done_bp       = True
                db = get_db()
                if db and st.session_state.selected_client_id:
                    try:
                        db.update_client(st.session_state.selected_client_id,
                                         {"buyer_persona": new_bp, "criterios": new_cr})
                        st.session_state.selected_client["buyer_persona"] = new_bp
                        st.session_state.selected_client["criterios"]     = new_cr
                    except: pass
                st.success(f"✅ Buyer Persona guardado para **{client_name}**. Continúa en Empresas →")

# ── TAB 3 · EMPRESAS ──────────────────────────────────────────────────────────
with tab3:
    if not st.session_state.done_bp:
        st.info("👈 Primero completa **ICP** y **Buyer Persona**.")
    else:
        st.subheader("Empresas objetivo")
        _n_excluidas = len(st.session_state.processed_domains or [])
        if _n_excluidas > 0:
            st.caption(f"🚫 {_n_excluidas} empresas ya prospectadas anteriormente serán excluidas automáticamente.")

        modo_empresas = st.radio(
            "¿Cómo quieres definir las empresas?",
            ["🤖 Recomendar empresas con IA", "📁 Subir Excel con empresas objetivo"],
            horizontal=True, key="modo_empresas",
        )

        if modo_empresas == "🤖 Recomendar empresas con IA":
            st.caption(f"Claude AI analiza la propuesta de valor, el sitio web y el ICP de **{client_name}** para recomendar empresas del mercado objetivo.")
            if st.button("🤖 Recomendar empresas", type="primary"):
                with st.spinner("Claude AI analizando el ICP..."):
                    try:
                        _dominios_rechazados = {r.get("dominio_web","") for r in st.session_state.empresas_rechazadas if r.get("dominio_web")}
                        # Exclusión permanente del cliente (empresas ya trabajadas)
                        _excl_perm = (st.session_state.selected_client or {}).get("exclusion_companies") or []
                        _excl_perm_dominios = {e.get("dominio_web","").lower() for e in _excl_perm if e.get("dominio_web")}
                        _excl_perm_nombres  = [e.get("nombre_empresa","") for e in _excl_perm if e.get("nombre_empresa")]
                        ya_vistos = list({e.get("dominio_web","") for e in st.session_state.empresas if e.get("dominio_web")} |
                                         set(st.session_state.processed_domains or []) |
                                         _dominios_rechazados |
                                         _excl_perm_dominios)
                        _razones_rech = [r.get("razon_rechazo","").strip() for r in st.session_state.empresas_rechazadas if r.get("razon_rechazo","").strip()]
                        _lookalike = (st.session_state.selected_client or {}).get("lookalike_companies") or None
                        empresas = get_company_recommendations(
                            st.session_state.icp, st.session_state.buyer_persona,
                            st.session_state.criterios,
                            n=st.session_state.n_empresas, demo=DEMO,
                            propuesta_de_valor=st.session_state.propuesta_de_valor,
                            excluir_dominios=ya_vistos,
                            excluir_nombres=_excl_perm_nombres or None,
                            razones_rechazo=_razones_rech or None,
                            lookalike_empresas=_lookalike)
                        st.session_state.empresas           = empresas
                        st.session_state.empresas_aprobadas = []
                        st.session_state.done_empresas      = False
                        # Persistir en Supabase para que sobrevivan refrescos y cambios de ICP/BP
                        try:
                            _db_persist = get_db()
                            if _db_persist and st.session_state.selected_client_id:
                                _db_persist.update_client(st.session_state.selected_client_id,
                                                          {"empresas_activas": empresas})
                        except Exception:
                            pass
                        try:
                            _db_log = get_db()
                            if _db_log and hasattr(_db_log, "log_usage"):
                                _db_log.log_usage(
                                    service="claude", action="recomendar_empresas",
                                    units=len(empresas),
                                    client_id=st.session_state.selected_client_id or "",
                                    client_name=client_name or "",
                                    user_username=st.session_state.get("auth_username",""),
                                    details={"n_empresas": len(empresas)},
                                )
                        except Exception:
                            pass
                    except Exception as e:
                        _err = str(e)
                        if "529" in _err or "overloaded" in _err.lower():
                            st.warning("⚠️ La IA está temporalmente sobrecargada. Espera unos segundos e intenta de nuevo.")
                        else:
                            st.error(f"Error: {e}")

        else:  # Subir Excel
            st.caption("Sube una lista de empresas en Excel o CSV. Columnas requeridas: **nombre_empresa**, **dominio_web**. Opcionales: pais, industria, tamano_empleados, razon_fit, linkedin_url.")
            uploaded = st.file_uploader("Selecciona el archivo", type=["xlsx","xls","csv"], key="upload_empresas")
            if uploaded:
                try:
                    import pandas as _pd_up
                    if uploaded.name.endswith(".csv"):
                        df_up = _pd_up.read_csv(uploaded)
                    else:
                        df_up = _pd_up.read_excel(uploaded)
                    # Normalizar nombres de columnas
                    df_up.columns = [c.strip().lower().replace(" ","_") for c in df_up.columns]
                    # Mapear columnas alternativas comunes
                    col_map = {
                        "empresa":"nombre_empresa","company":"nombre_empresa","name":"nombre_empresa",
                        "dominio":"dominio_web","domain":"dominio_web","website":"dominio_web","url":"dominio_web",
                        "pais":"pais","country":"pais","país":"pais",
                        "industria":"industria","industry":"industria","sector":"industria",
                        "empleados":"tamano_empleados","employees":"tamano_empleados","size":"tamano_empleados",
                        "razon":"razon_fit","reason":"razon_fit","fit":"razon_fit","notes":"razon_fit",
                        "linkedin":"linkedin_url",
                    }
                    df_up.rename(columns={k:v for k,v in col_map.items() if k in df_up.columns}, inplace=True)
                    if "nombre_empresa" not in df_up.columns:
                        st.error("El archivo debe tener una columna 'nombre_empresa' o 'empresa'.")
                    else:
                        # Normalizar columna "aprobada" si existe (SI/NO del cliente)
                        _aprobada_col = next((c for c in df_up.columns if "aprobada" in c.lower()), None)
                        empresas_up_raw = df_up.fillna("").to_dict("records")
                        empresas_up = []
                        _rechazadas_cliente = []
                        for e in empresas_up_raw:
                            e.setdefault("nombre_empresa", ""); e.setdefault("dominio_web","")
                            e.setdefault("pais",""); e.setdefault("industria","")
                            e.setdefault("tamano_empleados",""); e.setdefault("razon_fit","Wishlist del cliente")
                            e.setdefault("linkedin_url","")
                            if _aprobada_col:
                                _apro_val = str(e.get(_aprobada_col,"")).strip().upper()
                                e["aprobada"] = _apro_val in ("SI","SÍ","YES","S","Y","1","APROBADA","APROBADO","TRUE")
                                if not e["aprobada"]:
                                    _rechazadas_cliente.append(e)
                                    continue
                            else:
                                e["aprobada"] = True
                            empresas_up.append(e)

                        if _aprobada_col and _rechazadas_cliente:
                            st.info(f"ℹ️ Se filtraron **{len(_rechazadas_cliente)} empresas rechazadas** por el cliente (columna 'Aprobada = NO').")
                        st.success(f"✅ {len(empresas_up)} empresas aprobadas cargadas desde el archivo.")
                        _preview_cols = [c for c in ["nombre_empresa","dominio_web","pais","industria"] if c in _pd_up.DataFrame(empresas_up).columns]
                        if empresas_up:
                            st.dataframe(_pd_up.DataFrame(empresas_up)[_preview_cols],
                                         use_container_width=True, hide_index=True)
                        if st.button("✅ Usar estas empresas", type="primary", key="confirmar_upload"):
                            st.session_state.empresas           = empresas_up
                            st.session_state.empresas_aprobadas = empresas_up
                            st.session_state.done_empresas      = True
                            # Añadir empresas rechazadas por el cliente al registro
                            _dom_ya_rech = {r.get("dominio_web","") for r in st.session_state.empresas_rechazadas}
                            for _er_c in _rechazadas_cliente:
                                _dom_er_c = _er_c.get("dominio_web","")
                                if _dom_er_c and _dom_er_c not in _dom_ya_rech:
                                    st.session_state.empresas_rechazadas.append({
                                        "nombre_empresa": _er_c.get("nombre_empresa",""),
                                        "dominio_web"   : _dom_er_c,
                                        "razon_rechazo" : "Rechazada por el cliente",
                                        "fecha_rechazo" : datetime.now().strftime("%Y-%m-%d"),
                                    })
                            st.success("✅ Empresas cargadas. Continúa en Contactos →")
                            st.rerun()
                except Exception as e:
                    st.error(f"Error leyendo el archivo: {e}")

        if st.session_state.empresas:
            total = len(st.session_state.empresas)

            # ── Exportar para aprobación del cliente ──────────────────────────
            if modo_empresas == "🤖 Recomendar empresas con IA":
                import pandas as _pd_aprov
                _df_aprov = _pd_aprov.DataFrame([{
                    "Empresa"          : e.get("nombre_empresa",""),
                    "Dominio web"      : e.get("dominio_web",""),
                    "País"             : e.get("pais",""),
                    "Industria"        : e.get("industria",""),
                    "Empleados"        : e.get("tamano_empleados",""),
                    "Razón de fit"     : e.get("razon_fit",""),
                    "Aprobada (SI/NO)" : "",
                } for e in st.session_state.empresas])
                _buf_aprov = io.BytesIO()
                with _pd_aprov.ExcelWriter(_buf_aprov, engine="openpyxl") as _wr_aprov:
                    _df_aprov.to_excel(_wr_aprov, index=False, sheet_name="Empresas")
                _buf_aprov.seek(0)
                st.download_button(
                    "📤 Exportar para aprobación del cliente",
                    data=_buf_aprov.getvalue(),
                    file_name=f"empresas_para_aprobar_{client_name.replace(' ','_')}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    help="Descarga el Excel, envíaselo al cliente para que marque SI/NO en la columna 'Aprobada'. Luego sube el Excel aprobado en 'Subir Excel con empresas objetivo'.",
                    use_container_width=False,
                )
                st.caption("💡 El cliente marca **SI** o **NO** en la columna *Aprobada*. Luego sube ese Excel en **'📁 Subir Excel con empresas objetivo'**.")

            # ── Inicializar estado: todas aprobadas por defecto ───────────────
            def _clean(v):
                return str(v).strip() if v and str(v).strip() not in ("nan","None","") else ""

            for _ei, _emp in enumerate(st.session_state.empresas):
                _ekey = f"emp_estado_{_emp.get('dominio_web', _ei)}"
                if _ekey not in st.session_state:
                    # Por defecto aprobada (True), o respetar valor guardado
                    st.session_state[_ekey] = bool(_emp.get("aprobada", True))

            # ── Botones masivos ───────────────────────────────────────────────
            _ba, _br, _ = st.columns([2, 2, 6])
            if _ba.button("✅ Aprobar todas", key="aprobar_todas_emp", use_container_width=True):
                for _ei, _emp in enumerate(st.session_state.empresas):
                    st.session_state[f"emp_estado_{_emp.get('dominio_web', _ei)}"] = True
                st.rerun()
            if _br.button("❌ Rechazar todas", key="rechazar_todas_emp", use_container_width=True):
                for _ei, _emp in enumerate(st.session_state.empresas):
                    st.session_state[f"emp_estado_{_emp.get('dominio_web', _ei)}"] = False
                st.rerun()

            st.markdown("<hr style='margin:6px 0 10px 0'>", unsafe_allow_html=True)

            # ── Filas de empresas ─────────────────────────────────────────────
            for _ei, _emp in enumerate(st.session_state.empresas):
                _ekey  = f"emp_estado_{_emp.get('dominio_web', _ei)}"
                _apro  = st.session_state.get(_ekey, True)
                _nom   = _emp.get("nombre_empresa", "Sin nombre")
                _dom   = _emp.get("dominio_web", "")
                _url   = f"https://{_dom}" if _dom and not _dom.startswith("http") else _dom
                _pai   = _clean(_emp.get("pais", ""))
                _ind   = _clean(_emp.get("industria", ""))
                _tam   = _clean(_emp.get("tamano_empleados", ""))
                _rf    = _clean(_emp.get("razon_fit", ""))

                # Colores según estado
                _bg    = "#e8f5e9" if _apro else "#ffebee"
                _badge = "✅ Aprobada" if _apro else "❌ Rechazada"
                _badge_color = "#2e7d32" if _apro else "#c62828"

                st.markdown(f"""
<div style="background:{_bg};border-radius:10px;padding:14px 18px;margin-bottom:10px">
  <div style="display:flex;justify-content:space-between;align-items:flex-start;gap:12px">
    <div style="flex:0 0 auto">
      <span style="background:{_badge_color};color:white;font-size:0.75rem;
                   padding:2px 10px;border-radius:12px;font-weight:600">{_badge}</span>
    </div>
    <div style="flex:1 1 0;min-width:0">
      <div style="font-weight:700;font-size:1rem;color:#1a1a1a">{_nom}</div>
      <div style="font-size:0.82rem;color:#555;margin-top:2px">{_pai} &nbsp;·&nbsp; {_ind} &nbsp;·&nbsp; {_tam} empleados</div>
      <div style="font-size:0.87rem;color:#333;margin-top:6px;line-height:1.5">{_rf}</div>
    </div>
  </div>
</div>""", unsafe_allow_html=True)

                # Botones de acción debajo de la tarjeta
                _bc1, _bc2, _bc3, _bsite = st.columns([2, 2, 6, 1])
                if _bc1.button("✅ Aprobar", key=f"ap_e_{_ei}",
                               use_container_width=True, disabled=_apro):
                    st.session_state[_ekey] = True
                    st.rerun()
                if _bc2.button("❌ Rechazar", key=f"re_e_{_ei}",
                               use_container_width=True, disabled=not _apro):
                    st.session_state[_ekey] = False
                    st.rerun()
                if _url:
                    _bsite.link_button("🔗", _url, use_container_width=True)

                st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)

            st.markdown("<hr style='margin:10px 0'>", unsafe_allow_html=True)

            # ── Métricas y confirmar ──────────────────────────────────────────
            n_aprobadas  = sum(
                1 for _ei, _emp in enumerate(st.session_state.empresas)
                if st.session_state.get(f"emp_estado_{_emp.get('dominio_web', _ei)}", True)
            )
            n_rechazadas = total - n_aprobadas

            col_stats1, col_stats2, col_confirm = st.columns([1, 1, 2])
            col_stats1.metric("✅ Aprobadas",  n_aprobadas)
            col_stats2.metric("❌ Rechazadas", n_rechazadas)

            with col_confirm:
                if st.button("💾 Confirmar selección", type="primary", key="confirmar_empresas"):
                    empresas_actualizadas = []
                    for _ei, _emp in enumerate(st.session_state.empresas):
                        _ekey2 = f"emp_estado_{_emp.get('dominio_web', _ei)}"
                        _emp_c = _emp.copy()
                        _emp_c["aprobada"] = bool(st.session_state.get(_ekey2, True))
                        empresas_actualizadas.append(_emp_c)
                    st.session_state.empresas = empresas_actualizadas
                    st.session_state.empresas_aprobadas = [
                        e for e in empresas_actualizadas if e.get("aprobada", True)
                    ]
                    # ── Guardar rechazadas ─────────────────────────────────────
                    _rechazadas_nuevas = [e for e in empresas_actualizadas if not e.get("aprobada", True)]
                    _dominios_ya_rechazados = {r.get("dominio_web","") for r in st.session_state.empresas_rechazadas}
                    for _er in _rechazadas_nuevas:
                        _dom_er = _er.get("dominio_web","")
                        if _dom_er and _dom_er not in _dominios_ya_rechazados:
                            st.session_state.empresas_rechazadas.append({
                                "nombre_empresa" : _er.get("nombre_empresa",""),
                                "dominio_web"    : _dom_er,
                                "industria"      : _er.get("industria",""),
                                "pais"           : _er.get("pais",""),
                                "razon_fit"      : _er.get("razon_fit",""),
                                "razon_rechazo"  : "",
                                "fecha_rechazo"  : datetime.now().strftime("%Y-%m-%d"),
                            })
                    # Guardar rechazadas + empresas activas (persisten hasta marcar como prospectadas)
                    _db_conf = get_db()
                    if _db_conf and st.session_state.selected_client_id:
                        try:
                            _db_conf.update_client(st.session_state.selected_client_id, {
                                "empresas_rechazadas": st.session_state.empresas_rechazadas,
                                "empresas_activas"   : empresas_actualizadas,
                            })
                        except Exception as _e_conf:
                            st.warning(f"⚠️ No se pudo guardar en BD: {_e_conf}.")
                    st.session_state.done_empresas = True
                    st.success(f"✅ {n_aprobadas} empresas aprobadas. {len(_rechazadas_nuevas)} nuevas rechazadas registradas.")

            # ── Reemplazar rechazadas ─────────────────────────────────────────
            if n_rechazadas > 0:
                st.markdown("---")
                if DEMO:
                    st.info(f"⚠️ En Modo Demo no se pueden generar reemplazos. "
                            f"Desactiva el Modo Demo para buscar {n_rechazadas} empresa{'s' if n_rechazadas != 1 else ''} alternativa{'s' if n_rechazadas != 1 else ''}.",
                            icon="🔄")
                else:
                    if st.button(f"🔄 Reemplazar {n_rechazadas} rechazada{'s' if n_rechazadas != 1 else ''} por nuevas empresas",
                                 key="reemplazar_rechazadas"):
                        # Leer estado actual desde session_state
                        empresas_actualizadas2 = []
                        for _ei2, _emp2 in enumerate(st.session_state.empresas):
                            _ekey3 = f"emp_estado_{_emp2.get('dominio_web', _ei2)}"
                            _emp2c = _emp2.copy()
                            _emp2c["aprobada"] = bool(st.session_state.get(_ekey3, True))
                            empresas_actualizadas2.append(_emp2c)

                        # Dominios ya vistos (incluye rechazadas)
                        _dom_rech2 = {r.get("dominio_web","") for r in st.session_state.empresas_rechazadas if r.get("dominio_web")}
                        ya_vistos = list({e.get("dominio_web","") for e in empresas_actualizadas2 if e.get("dominio_web")} |
                                         set(st.session_state.processed_domains or []) | _dom_rech2)

                        with st.spinner(f"Buscando {n_rechazadas} empresas alternativas…"):
                            try:
                                _lookalike2 = (st.session_state.selected_client or {}).get("lookalike_companies") or None
                                nuevas = get_company_recommendations(
                                    st.session_state.icp,
                                    st.session_state.buyer_persona,
                                    st.session_state.criterios,
                                    n=n_rechazadas,
                                    demo=False,
                                    propuesta_de_valor=st.session_state.propuesta_de_valor,
                                    excluir_dominios=ya_vistos,
                                    lookalike_empresas=_lookalike2,
                                )
                                for e in nuevas:
                                    e["aprobada"] = True
                                aprobadas_actuales = [e for e in empresas_actualizadas2 if e.get("aprobada", True)]
                                st.session_state.empresas = aprobadas_actuales + nuevas
                                st.session_state.empresas_aprobadas = []
                                st.session_state.done_empresas = False
                                st.rerun()
                            except Exception as e:
                                st.error(f"Error al buscar reemplazos: {e}")

# ── TAB 4 · CONTACTOS ─────────────────────────────────────────────────────────
with tab4:
    if not st.session_state.done_empresas:
        st.info("👈 Primero genera y aprueba las **empresas**.")
    else:
        st.subheader("Búsqueda de contactos")
        _aprobadas = st.session_state.empresas_aprobadas or []
        st.caption(f"**{len(_aprobadas)} empresas aprobadas** avanzan a esta etapa.")

        import pandas as _pd_c
        import io as _io_c

        # ── URL de Sales Navigator ────────────────────────────────────────────
        st.markdown("#### 🔗 URL de búsqueda para LinkedIn Sales Navigator")
        bp   = st.session_state.buyer_persona or {}
        # Si session_state.icp está vacío, intentar cargarlo del cliente seleccionado
        if not st.session_state.icp and st.session_state.selected_client:
            st.session_state.icp = st.session_state.selected_client.get("icp") or {}
        if not st.session_state.buyer_persona and st.session_state.selected_client:
            st.session_state.buyer_persona = st.session_state.selected_client.get("buyer_persona") or {}
            bp = st.session_state.buyer_persona or {}
        icp_ = st.session_state.icp or {}
        # Calcular cuántos lotes necesitamos (máx _BATCH_SIZE empresas por URL)
        _total_aprob  = len(_aprobadas)
        _n_batches    = max(1, -(-_total_aprob // _BATCH_SIZE))  # ceil division
        _snav_batches = [
            generar_url_sales_navigator(bp, icp_, empresas_aprobadas=_aprobadas,
                                        _company_offset=i * _BATCH_SIZE)
            for i in range(_n_batches)
        ]
        snav_url = _snav_batches[0]  # compatibilidad con código que usa snav_url más abajo

        # Mostrar URL(s) generadas
        with st.expander("🔗 Ver URL de Sales Navigator generada", expanded=True):
            if _n_batches == 1:
                st.text_area(
                    "URL con empresas aprobadas + filtros de Buyer Persona:",
                    value=snav_url, height=120,
                )
                st.link_button("🚀 Abrir búsqueda en Sales Navigator", url=snav_url,
                               type="primary",
                               help="Abre la búsqueda en LinkedIn Sales Navigator en una nueva pestaña")
            else:
                st.info(f"📋 {_total_aprob} empresas → divididas en **{_n_batches} búsquedas** de máx {_BATCH_SIZE} empresas c/u.")
                for _bi, _burl in enumerate(_snav_batches):
                    _start = _bi * _BATCH_SIZE + 1
                    _end   = min((_bi + 1) * _BATCH_SIZE, _total_aprob)
                    _bcols = st.columns([3, 1])
                    with _bcols[0]:
                        st.text_area(f"Búsqueda {_bi+1} (empresas {_start}–{_end}):",
                                     value=_burl, height=100, key=f"snav_url_{_bi}")
                    with _bcols[1]:
                        st.link_button(f"🚀 Abrir búsqueda {_bi+1}",
                                       url=_burl, use_container_width=True)

            cargos_lista    = bp.get("cargos_objetivo",  [])
            excluidos_lista = bp.get("cargos_excluidos", [])
            if cargos_lista:
                _kw_preview = " OR ".join(cargos_lista[:8]) + ("…" if len(cargos_lista) > 8 else "")
                if excluidos_lista:
                    _ex_preview = " NOT ".join(excluidos_lista[:6])
                    st.caption(f"🔑 Keywords: {_kw_preview} NOT {_ex_preview}")
                else:
                    st.caption(f"🔑 Keywords: {_kw_preview}")
            n_emp_url = len(_aprobadas)
            st.caption(f"🏢 Empresas ({n_emp_url}): {', '.join([e.get('nombre_empresa','') for e in _aprobadas[:4]])}{'…' if n_emp_url>4 else ''}")
            # Mostrar países que se incluyen en el filtro de geografía
            _icp_paises_url = (icp_ or {}).get("geografias", [])
            if isinstance(_icp_paises_url, list) and _icp_paises_url:
                st.caption(f"🌍 Países en filtro: {', '.join(_icp_paises_url[:6])}{'…' if len(_icp_paises_url)>6 else ''}")
            else:
                st.warning("⚠️ Sin filtro de países — el ICP no tiene geografías o la sesión se reinició. "
                           "Ve a la pestaña **ICP** y vuelve a guardar para aplicar el filtro de países.")

        st.divider()

        # ── Flujo manual con Lemlist plugin ──────────────────────────────────
        st.markdown("#### 📋 Flujo de importación de contactos")

        if DEMO:
            st.info("⚠️ Modo Demo — contactos ficticios.")
            if st.button("▶ Generar contactos demo", type="primary"):
                st.session_state.contactos_clay             = _demo_contacts()
                st.session_state.contactos_aprobados        = []
                st.session_state.contacts_pushed_to_enrich  = False
                st.session_state.done_clay                  = False
                st.rerun()
        else:
            st.info(
                "**Flujo recomendado para importar contactos a Lemlist:**\n\n"
                "1. 🚀 Haz clic en **«Abrir búsqueda en Sales Navigator»** (arriba) para revisar los leads\n"
                "2. 🔌 Usa el **plugin de Lemlist** en el navegador para importar los leads a la lista "
                "**«Contactos App Prospección (Por enriquecer)»** — sin activar enrichment aún\n"
                "3. 📋 En Lemlist, filtra los perfiles fit y muévelos a la lista "
                "**«Contactos Validados para enriquecer»**\n"
                "4. ➕ En Lemlist, agrega esos contactos a la campaña **«Validados para enriquecer»** "
                "— esto es necesario para que la API pueda leerlos\n"
                "5. ✉️ En Lemlist, selecciona todos los contactos de esa campaña y activa el enrichment "
                "(email + teléfono) manualmente\n"
                "6. 📱 Vuelve aquí a la pestaña **Enriquecimiento** para completar los teléfonos faltantes con Lusha"
            )


# ── TAB 5 · ENRIQUECIMIENTO ───────────────────────────────────────────────────
with tab5:
    st.subheader("📱 Enriquecimiento de teléfonos con Lusha")

    lm_key_tab5 = lemlist_key()
    if not lm_key_tab5:
        st.warning(
            "⚠️ No hay API Key de Lemlist configurada para este cliente.\n\n"
            "**Cómo solucionarlo:**\n"
            "1. Ve a ⚙️ **Gestión de clientes** (panel izquierdo)\n"
            "2. Edita este cliente\n"
            "3. Pega la API Key desde **Lemlist → Settings → Integrations → API**\n"
            "4. Guarda y vuelve a esta pestaña"
        )
    elif not LUSHA_API_KEY:
        st.error("⚠️ Falta LUSHA_API_KEY en Streamlit Secrets.")
    else:
        # ── Recordatorio del flujo ─────────────────────────────────────────
        with st.expander("ℹ️ ¿Cómo funciona este flujo?", expanded=False):
            st.markdown(
                "**Antes de usar esta sección, asegúrate de haber completado:**\n\n"
                "1. ✅ Importaste los leads a Lemlist con el plugin desde Sales Navigator "
                "(lista «Contactos App Prospección (Por enriquecer)»)\n"
                "2. ✅ Filtraste los perfiles fit y los moviste a la lista "
                "**«Contactos Validados para enriquecer»**\n"
                "3. ✅ En Lemlist, seleccionaste todos los contactos de esa lista y activaste "
                "el enrichment manual (email + teléfono) con la waterfall de Lemlist\n\n"
                "**Lusha completa solo lo que Lemlist no encontró (teléfonos faltantes).**"
            )

        st.divider()

        # ── Paso 1: Seleccionar lista de Lemlist ───────────────────────────
        st.markdown("#### 📋 Paso 1 · Seleccionar lista en Lemlist")

        _lm5 = LemlistClient(lm_key_tab5)
        _NOMBRE_LISTA_VALIDADOS = "Contactos Validados para enriquecer"


        # Nota: La API de Lemlist NO permite listar contactos por lista (/contacts siempre
        # requiere idsOrEmails). Solo las CAMPAÑAS exponen sus leads por API.
        st.info(
            "ℹ️ **Nota técnica de Lemlist:** La API de Lemlist solo permite leer contactos "
            "de **campañas** (secuencias), no de listas de contactos.\n\n"
            "**Solución rápida:** En Lemlist, agrega los contactos de tu lista a una campaña "
            "(puede ser una campaña \"dummy\" solo para este fin) → selecciónala abajo.",
            icon="💡"
        )

        # Almacenar contactos cargados en session_state para persistencia
        if "enrich_contacts_loaded" not in st.session_state:
            st.session_state.enrich_contacts_loaded = []
        if "enrich_lista_id" not in st.session_state:
            st.session_state.enrich_lista_id = ""
        if "enrich_lista_nombre_activa" not in st.session_state:
            st.session_state.enrich_lista_nombre_activa = ""
        if "enrich_all_camps" not in st.session_state:
            st.session_state.enrich_all_camps = []

        # Cargar campañas disponibles
        if not st.session_state.enrich_all_camps:
            with st.spinner("Cargando campañas de Lemlist…"):
                try:
                    st.session_state.enrich_all_camps = _lm5.get_campaigns()
                except Exception as _ec:
                    st.error(f"❌ Error al conectar con Lemlist: {_ec}")

        _all_camps5 = st.session_state.enrich_all_camps
        if _all_camps5:
            _camp_names5 = [c.get("name","") for c in _all_camps5]
            col_lista, col_btn_lista, col_refresh = st.columns([3, 1, 1])
            with col_lista:
                _sel_camp_name = st.selectbox(
                    "Selecciona la campaña en Lemlist:",
                    options=_camp_names5,
                    key="enrich_lista_nombre",
                    help="Selecciona la campaña que contiene los contactos a enriquecer con Lusha",
                )
            with col_btn_lista:
                st.write("")
                st.write("")
                _btn_cargar = st.button("🔄 Cargar contactos", type="primary", key="btn_cargar_lista_enrich")
            with col_refresh:
                st.write("")
                st.write("")
                if st.button("↺ Actualizar", key="btn_refresh_listas"):
                    st.session_state.enrich_all_camps = []
                    st.rerun()

            if _btn_cargar and _sel_camp_name:
                _match5 = next((c for c in _all_camps5 if c.get("name","") == _sel_camp_name), None)
                if _match5:
                    with st.spinner(f"Cargando contactos de «{_sel_camp_name}»…"):
                        try:
                            _camp_id_5 = _match5.get("_id") or _match5.get("id","")
                            # get_campaign_contacts = leads → contactIds → detalles completos
                            _leads_raw = _lm5.get_campaign_contacts(_camp_id_5, limit=1000)
                            _contacts_norm = []
                            for lead in _leads_raw:
                                # Lemlist devuelve datos en lead["fields"] (nivel anidado)
                                _f  = lead.get("fields") or {}
                                _fn = _f.get("firstName") or lead.get("firstName") or lead.get("first_name") or ""
                                _ln = _f.get("lastName")  or lead.get("lastName")  or lead.get("last_name")  or ""
                                # Teléfono: Lemlist guarda phone1/phone2 en fields
                                _phone_val = (
                                    _f.get("phone") or _f.get("phoneNumber") or
                                    _f.get("phone1") or _f.get("phone2") or
                                    lead.get("phone") or ""
                                )
                                # LinkedIn: URL Sales Nav en raíz como linkedinUrlSalesNav
                                _li_val = (
                                    lead.get("linkedinUrlSalesNav") or
                                    lead.get("linkedinUrl") or lead.get("linkedInUrl") or
                                    _f.get("linkedinUrl") or _f.get("linkedInUrl") or
                                    _f.get("linkedin_url") or _f.get("linkedin") or
                                    _f.get("linkedinSalesUrl") or _f.get("salesNavigatorUrl") or ""
                                )
                                # leadId (lea_...) necesario para actualizar variables en Lemlist
                                # get_campaign_contacts lo inyecta en _leadId desde el mapa contactId→leadId
                                _camps = lead.get("campaigns") or []
                                _lead_id = (
                                    lead.get("_leadId") or                          # inyectado por get_campaign_contacts
                                    (_camps[0].get("leadId") if _camps else None) or  # fallback: campaigns array
                                    ""
                                )
                                # campaign_id para PATCH /campaigns/{cid}/leads/{lid}
                                _camp_id_lead = (
                                    (_camps[0].get("campaignId") if _camps else None) or
                                    _camp_id_5 or ""
                                )
                                _job   = (_f.get("jobTitle") or _f.get("job_title") or _f.get("title") or _f.get("position") or
                                          lead.get("jobTitle") or lead.get("job_title") or lead.get("title") or lead.get("position") or "")
                                _comp  = (_f.get("companyName") or _f.get("company") or _f.get("company_name") or
                                          lead.get("companyName") or lead.get("company") or lead.get("company_name") or "")
                                _ctry  = (_f.get("country") or _f.get("location") or
                                          lead.get("country") or lead.get("location") or "")
                                _contacts_norm.append({
                                    "first_name"  : _fn,
                                    "last_name"   : _ln,
                                    "full_name"   : f"{_fn} {_ln}".strip() or lead.get("fullName") or "",
                                    "job_title"   : _job,
                                    "company_name": _comp,
                                    "email"       : lead.get("email") or _f.get("email") or "",
                                    "phone"       : _phone_val,
                                    "phone_source": "Lemlist" if _phone_val else "",
                                    "linkedin_url": _li_val,
                                    "country"     : _ctry,
                                    "lead_id"     : _lead_id,
                                    "campaign_id" : _camp_id_lead,
                                })
                            # Guardar raw de primer contacto para debug LinkedIn URL
                            if _leads_raw:
                                st.session_state["_debug_lemlist_raw_lead"] = _leads_raw[0]
                            st.session_state.enrich_contacts_loaded     = _contacts_norm
                            st.session_state.enrich_lista_id            = _camp_id_5
                            st.session_state.enrich_lista_nombre_activa = _sel_camp_name
                            st.rerun()
                        except Exception as _e5:
                            st.error(f"❌ Error al cargar contactos: {_e5}")

        # ── Mostrar contactos cargados ─────────────────────────────────────
        _contacts5 = st.session_state.enrich_contacts_loaded
        _lista_id5  = st.session_state.enrich_lista_id
        _lista_nom5 = st.session_state.enrich_lista_nombre_activa


        if not _contacts5:
            st.info("👆 Ingresa el nombre de la lista y haz clic en **«Cargar contactos»** para comenzar.")
        else:
            st.success(f"✅ Lista cargada: **{_lista_nom5}** · {len(_contacts5)} contactos")

            # Métricas de estado
            _total5    = len(_contacts5)
            _con_email = sum(1 for c in _contacts5 if c.get("email","").strip())
            _con_tel5  = sum(1 for c in _contacts5 if c.get("phone","").strip())
            _sin_tel5  = _total5 - _con_tel5
            _sin_email_sin_tel = sum(
                1 for c in _contacts5
                if not c.get("email","").strip() and not c.get("phone","").strip()
            )

            s1, s2, s3, s4 = st.columns(4)
            s1.metric("Total contactos",    _total5)
            s2.metric("✉️ Con email",        _con_email,
                      help="Lemlist encontró el email")
            s3.metric("📞 Con teléfono",     _con_tel5,
                      help="Ya tienen teléfono (Lemlist o manual)")
            s4.metric("📵 Sin teléfono",     _sin_tel5,
                      delta=f"-{_sin_tel5}" if _sin_tel5 else None,
                      delta_color="inverse",
                      help="Lusha buscará teléfono para estos contactos")

            if _sin_email_sin_tel > 0:
                st.warning(
                    f"⚠️ **{_sin_email_sin_tel} contactos no tienen email ni teléfono.** "
                    "Esto puede significar que el enrichment de Lemlist aún no ha corrido. "
                    "Lusha igualmente intentará buscarles teléfono si confirmas que el enrichment ya corrió."
                )

            st.divider()

            # ── LinkedIn URLs manuales (para contactos sin URL en Lemlist) ──
            if _sin_tel5 > 0:
                _contactos_sin_tel_prev = [c for c in _contacts5 if not c.get("phone","").strip()]
                _sin_li = [c for c in _contactos_sin_tel_prev if not c.get("linkedin_url","").strip()]
                if _sin_li:
                    with st.expander(
                        f"🔗 **Pegar URLs de LinkedIn** para los {len(_sin_li)} contactos sin URL "
                        f"(mejora mucho la precisión de Lusha)", expanded=False
                    ):
                        st.caption(
                            "Lusha encuentra teléfonos con mucha más precisión cuando tiene la URL de LinkedIn. "
                            "Pega aquí la URL de perfil de cada contacto antes de correr Lusha."
                        )
                        # Inicializar dict de LinkedIn overrides en session_state
                        if "lusha_linkedin_overrides" not in st.session_state:
                            st.session_state["lusha_linkedin_overrides"] = {}
                        for _sc in _sin_li:
                            _key_li = _sc.get("email","") or _sc.get("full_name","")
                            _cur_li = st.session_state["lusha_linkedin_overrides"].get(_key_li, "")
                            _new_li = st.text_input(
                                f"{_sc.get('full_name','')} ({_sc.get('company_name','')})",
                                value=_cur_li,
                                placeholder="https://www.linkedin.com/in/...",
                                key=f"li_override_{_key_li}",
                            )
                            if _new_li.strip():
                                st.session_state["lusha_linkedin_overrides"][_key_li] = _new_li.strip()

                # Aplicar overrides de LinkedIn al listado de contactos en memoria
                _li_overrides = st.session_state.get("lusha_linkedin_overrides", {})
                if _li_overrides:
                    for _c5 in _contacts5:
                        _k5 = _c5.get("email","") or _c5.get("full_name","")
                        if _k5 in _li_overrides and not _c5.get("linkedin_url","").strip():
                            _c5["linkedin_url"] = _li_overrides[_k5]

            # ── Paso 2: Confirmación y Lusha ──────────────────────────────
            st.markdown("#### 📞 Paso 2 · Complementar teléfonos faltantes con Lusha")

            if _sin_tel5 == 0:
                st.success("✅ ¡Todos los contactos ya tienen teléfono! No es necesario usar Lusha.")
            else:
                st.markdown(
                    f"Lusha buscará teléfono para los **{_sin_tel5} contactos** que no tienen número. "
                    "Esto incluye tanto los que tienen email como los que no."
                )

                # Checkbox de confirmación — el gate de timing
                _confirmado = st.checkbox(
                    "✅ Confirmo que ya corrí el enrichment en Lemlist para todos los contactos de esta lista "
                    "(email + teléfono). Entiendo que Lusha complementará solo los teléfonos faltantes.",
                    key="confirm_lemlist_enrich_done",
                )

                if not _confirmado:
                    st.info(
                        "☝️ Marca el checkbox cuando hayas completado el enrichment en Lemlist. "
                        "Así evitamos usar créditos de Lusha en contactos que aún no han sido procesados."
                    )
                else:
                    _contactos_sin_tel = [c for c in _contacts5 if not c.get("phone","").strip()]
                    _con_li_override   = sum(1 for c in _contactos_sin_tel if c.get("linkedin_url","").strip())
                    st.caption(
                        f"Se enviarán **{len(_contactos_sin_tel)} contactos** a Lusha. "
                        f"Los {_con_tel5} que ya tienen teléfono serán omitidos."
                        + (f" **{_con_li_override}** tienen URL de LinkedIn ✅" if _con_li_override else
                           " ⚠️ Ninguno tiene URL de LinkedIn — se buscará solo por nombre y empresa.")
                    )

                    if st.button(
                        f"🔍 Buscar {len(_contactos_sin_tel)} teléfonos faltantes con Lusha",
                        type="primary",
                        key="btn_lusha_enrich_tab5",
                    ):
                        # Separar contactos: con teléfono (no necesitan Lusha) y sin teléfono
                        _ya_con_tel   = [c for c in _contacts5 if c.get("phone","").strip()]
                        _necesitan_tel = [c for c in _contacts5 if not c.get("phone","").strip()]
                        _total_lusha  = len(_necesitan_tel)

                        pb_l5 = st.progress(0, text=f"Consultando Lusha (0/{_total_lusha})…")
                        _resultados_lusha = [c.copy() for c in _ya_con_tel]  # ya tienen tel, pasan directo
                        _encontrados = 0
                        _lusha_log = []   # log para debug

                        for i, c in enumerate(_necesitan_tel):
                            n = c.copy()
                            phone_l, email_l = _lusha_person(
                                n.get("first_name",""), n.get("last_name",""),
                                n.get("company_name",""), n.get("linkedin_url",""),
                                key=LUSHA_API_KEY, demo=DEMO
                            )
                            _lusha_log.append({
                                "nombre"   : n.get("full_name") or n.get("email",""),
                                "tel_lusha": phone_l or "—",
                                "email_lusha": email_l or "—",
                            })
                            if phone_l:
                                n["phone"]        = phone_l
                                n["phone_source"] = "Lusha"
                                _encontrados += 1
                            if not n.get("email","").strip() and email_l:
                                n["email"]        = email_l
                                n["email_source"] = "Lusha"
                            _resultados_lusha.append(n)
                            pb_l5.progress(
                                (i + 1) / _total_lusha,
                                text=f"Lusha {i+1}/{_total_lusha}: {n.get('full_name','') or n.get('email','')}"
                            )
                            if i < _total_lusha - 1:   # no esperar después del último
                                time.sleep(LUSHA_RATE_LIMIT)

                        pb_l5.empty()
                        st.session_state.enrich_contacts_loaded  = _resultados_lusha
                        st.session_state["lusha_run_log"]        = _lusha_log
                        st.session_state["lusha_run_encontrados"] = _encontrados
                        try:
                            _db_lusha_log = get_db()
                            if _db_lusha_log and _encontrados > 0 and hasattr(_db_lusha_log, "log_usage"):
                                _db_lusha_log.log_usage(
                                    service="lusha", action="enriquecer_telefonos",
                                    units=_encontrados,
                                    client_id=st.session_state.selected_client_id or "",
                                    client_name=(st.session_state.selected_client or {}).get("name",""),
                                    user_username=st.session_state.get("auth_username",""),
                                    details={"buscados": len(_contactos_sin_tel), "encontrados": _encontrados},
                                )
                        except Exception:
                            pass

                        st.session_state["lusha_lemlist_updated"] = 0  # se actualiza en Paso 3
                        st.rerun()

            # ── Resultado Lusha ───────────────────────────────────────────
            if st.session_state.get("lusha_run_log"):
                _enc5 = st.session_state.get("lusha_run_encontrados", 0)
                _log5 = st.session_state["lusha_run_log"]
                st.divider()
                if _enc5 > 0:
                    st.success(f"✅ Lusha encontró teléfono para **{_enc5}/{len(_log5)}** contactos.")
                else:
                    st.warning(f"⚠️ Lusha no encontró teléfono para ninguno de los {len(_log5)} contactos buscados.")
                import pandas as _pd_lusha
                st.dataframe(
                    _pd_lusha.DataFrame(_log5).rename(columns={
                        "nombre": "Contacto", "tel_lusha": "Teléfono encontrado", "email_lusha": "Email encontrado"
                    }),
                    use_container_width=True, hide_index=True
                )

            # ── Exportar a HubSpot ────────────────────────────────────────────
            if _contacts5:
                st.divider()
                st.markdown("#### 📥 Paso 3 · Exportar base completa a HubSpot")

                _empresas_exp  = st.session_state.empresas_aprobadas or st.session_state.empresas or []
                _total_exp     = len(_contacts5)
                _con_email_exp = sum(1 for c in _contacts5 if c.get("email","").strip())
                _con_tel_exp   = sum(1 for c in _contacts5 if c.get("phone","").strip())

                _ex1, _ex2, _ex3 = st.columns(3)
                _ex1.metric("Contactos",   _total_exp)
                _ex2.metric("Con email",   _con_email_exp)
                _ex3.metric("Con teléfono", _con_tel_exp)

                st.caption(
                    "El Excel incluye una hoja **Contactos** (formato HubSpot import) con teléfono unificado "
                    "de Lemlist + Lusha, y una hoja **Empresas** con toda la info del prospecting."
                )

                if st.button("📊 Generar Excel HubSpot", type="primary", key="btn_gen_excel_hs"):
                    with st.spinner("Generando Excel…"):
                        try:
                            _particularidades_exp = (st.session_state.selected_client or {}).get("particularidades_prospeccion","")
                            _excel_hs_bytes = build_excel_hubspot(
                                _contacts5, _empresas_exp, client_name,
                                particularidades=_particularidades_exp,
                            )
                            _fname_hs = (
                                f"hubspot_{client_name.replace(' ','_')}_"
                                f"{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
                            )
                            st.session_state["_hs_excel_bytes"] = _excel_hs_bytes
                            st.session_state["_hs_excel_fname"] = _fname_hs
                            try:
                                _db_exp_log = get_db()
                                if _db_exp_log and hasattr(_db_exp_log, "log_usage"):
                                    _db_exp_log.log_usage(
                                        service="export", action="generar_excel_hubspot",
                                        units=len(_contacts5),
                                        client_id=st.session_state.selected_client_id or "",
                                        client_name=(st.session_state.selected_client or {}).get("name",""),
                                        user_username=st.session_state.get("auth_username",""),
                                        details={"contactos": len(_contacts5), "empresas": len(_empresas_exp)},
                                    )
                            except Exception:
                                pass

                            # Guardar registro en historial (en el cliente de Supabase)
                            _hs_record = {
                                "fecha"    : datetime.now().isoformat(),
                                "filename" : _fname_hs,
                                "total"    : _total_exp,
                                "con_email": _con_email_exp,
                                "con_tel"  : _con_tel_exp,
                            }
                            _cur_hist = list(
                                (st.session_state.selected_client or {}).get("export_history", []) or []
                            )
                            _cur_hist.insert(0, _hs_record)
                            _db_hs = get_db()
                            if _db_hs and st.session_state.selected_client_id:
                                try:
                                    _db_hs.update_client(
                                        st.session_state.selected_client_id,
                                        {"export_history": _cur_hist}
                                    )
                                    if st.session_state.selected_client:
                                        st.session_state.selected_client["export_history"] = _cur_hist
                                except Exception:
                                    pass
                            st.rerun()
                        except Exception as _e_hs:
                            st.error(f"Error generando Excel: {_e_hs}")

                if st.session_state.get("_hs_excel_bytes"):
                    st.download_button(
                        "⬇️ Descargar Excel HubSpot",
                        data=st.session_state["_hs_excel_bytes"],
                        file_name=st.session_state.get("_hs_excel_fname", "hubspot_export.xlsx"),
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary",
                        use_container_width=True,
                        key="dl_hs_excel",
                    )

                # ── Historial de exportaciones ─────────────────────────────────
                _exp_hist = (st.session_state.selected_client or {}).get("export_history") or []
                if _exp_hist:
                    st.markdown("##### 📋 Historial de exportaciones")
                    import pandas as _pd_hs
                    _hist_rows = [{
                        "Fecha"      : h.get("fecha","")[:16].replace("T"," "),
                        "Archivo"    : h.get("filename",""),
                        "Contactos"  : h.get("total",0),
                        "Con email"  : h.get("con_email",0),
                        "Con teléfono": h.get("con_tel",0),
                    } for h in _exp_hist]
                    st.dataframe(
                        _pd_hs.DataFrame(_hist_rows),
                        use_container_width=True,
                        hide_index=True,
                    )

# ── TAB 6 · RESULTADOS ────────────────────────────────────────────────────────
with tab6:
    # Fuente de datos: Lemlist+Lusha (Tab 5) o Clay (flujo antiguo)
    _final6 = st.session_state.enrich_contacts_loaded or st.session_state.contactos_final or []

    if not _final6:
        st.info("👈 Carga los contactos en la pestaña **Enriquecimiento** para ver los resultados aquí.")
    else:
        _total6 = len(_final6)
        _ce6    = sum(1 for c in _final6 if c.get("email","").strip())
        _ct6    = sum(1 for c in _final6 if c.get("phone","").strip())
        _cl6    = sum(1 for c in _final6 if c.get("phone_source") == "Lusha")

        st.subheader(f"Prospección en curso — {client_name}")
        _r1, _r2, _r3, _r4 = st.columns(4)
        _r1.metric("Contactos",    _total6)
        _r2.metric("Con email",    _ce6,  f"{_ce6/_total6*100:.0f}%" if _total6 else "")
        _r3.metric("Con teléfono", _ct6,  f"{_ct6/_total6*100:.0f}%" if _total6 else "")
        _r4.metric("📞 Lusha",     _cl6,  help="Teléfonos encontrados por Lusha")

        _rows6 = [{
            "Nombre"   : norm_person(c.get("full_name",""), c.get("first_name",""), c.get("last_name","")),
            "Empresa"  : norm_company(c.get("company_name","")),
            "Cargo"    : c.get("job_title",""),
            "Email"    : c.get("email",""),
            "Teléfono" : c.get("phone",""),
            "País"     : c.get("country",""),
            "Fuente Tel": c.get("phone_source",""),
        } for c in _final6]
        st.dataframe(_pd.DataFrame(_rows6), use_container_width=True, hide_index=True)

        st.divider()
        _col_save6, _col_dl6 = st.columns(2)

        with _col_save6:
            if st.button("💾 Guardar en historial", type="primary", use_container_width=True):
                _db6 = get_db()
                if _db6 and st.session_state.selected_client_id:
                    try:
                        _stats6 = {
                            "total"  : _total6,
                            "emails" : _ce6,
                            "phones" : _ct6,
                            "lusha"  : _cl6,
                            "client" : client_name,
                            "date"   : datetime.now().isoformat(),
                        }
                        _stats6["user"] = st.session_state.get("auth_username", "")
                        _db6.save_run(
                            st.session_state.selected_client_id,
                            st.session_state.empresas_aprobadas or st.session_state.empresas,
                            _final6,
                            _stats6,
                        )
                        try:
                            if hasattr(_db6, "log_usage"):
                                _db6.log_usage(
                                    service="export", action="guardar_historial",
                                    client_id=st.session_state.selected_client_id,
                                    client_name=client_name,
                                    user_username=st.session_state.get("auth_username",""),
                                    details={"total": _total6, "emails": _ce6, "phones": _ct6, "lusha": _cl6},
                                )
                        except Exception:
                            pass
                        st.success("✅ Guardado en historial")
                    except Exception as _e6:
                        st.error(f"Error: {_e6}")
                else:
                    st.warning("Supabase no configurado.")

        with _col_dl6:
            pass

        # ── Marcar empresas como ya prospectadas ──────────────────────────────
        st.divider()
        _empresas_para_marcar = st.session_state.empresas_aprobadas or st.session_state.empresas
        _n_para_marcar = len([e for e in _empresas_para_marcar if e.get("dominio_web")])
        if _n_para_marcar > 0:
            st.markdown("#### ✅ ¿Terminaste la prospección de estas empresas?")
            st.caption(
                f"Al marcar las **{_n_para_marcar} empresas** como ya prospectadas, "
                "la IA no las volverá a recomendar en futuras búsquedas."
            )
            if st.button(f"📌 Marcar {_n_para_marcar} empresa{'s' if _n_para_marcar != 1 else ''} como ya prospectadas",
                         key="marcar_prospectadas", use_container_width=True):
                _db_mark = get_db()
                _dominios_nuevos = {e.get("dominio_web","").lower() for e in _empresas_para_marcar if e.get("dominio_web")}
                _todos_procesados = list(
                    set(st.session_state.processed_domains or []) | _dominios_nuevos
                )
                st.session_state.processed_domains = _todos_procesados
                if _db_mark and st.session_state.selected_client_id:
                    try:
                        _db_mark.update_client(st.session_state.selected_client_id, {
                            "processed_domains": _todos_procesados,
                            "empresas_activas" : [],   # limpiar para próximo ciclo
                        })
                        st.session_state.empresas           = []
                        st.session_state.empresas_aprobadas = []
                        st.session_state.done_empresas      = False
                        st.success(f"✅ {_n_para_marcar} empresas marcadas como ya prospectadas. La IA las excluirá en futuras búsquedas.")
                    except Exception as _e_mark:
                        st.error(f"Error al guardar: {_e_mark}")
                else:
                    st.success(f"✅ {_n_para_marcar} empresas marcadas como ya prospectadas (solo en sesión).")

# ── TAB 7 · HISTORIAL ────────────────────────────────────────────────────────
with tab7:
    st.subheader(f"Historial de prospecciones — {client_name}")
    _db7 = get_db()
    if not _db7:
        st.info("Conecta Supabase para ver el historial.")
    elif not st.session_state.selected_client_id:
        st.info("Selecciona un cliente para ver su historial.")
    else:
        try:
            _runs7 = _db7.get_runs(st.session_state.selected_client_id)
            if not _runs7:
                st.info("No hay prospecciones guardadas aún. Ve a **Resultados** y haz clic en **💾 Guardar en historial**.")
            else:
                # ── Totales acumulados ─────────────────────────────────────
                _tot_runs    = len(_runs7)
                _tot_leads   = sum((r.get("stats") or {}).get("total",0)  for r in _runs7)
                _tot_emails  = sum((r.get("stats") or {}).get("emails",0) for r in _runs7)
                _tot_phones  = sum((r.get("stats") or {}).get("phones",0) for r in _runs7)

                st.caption(f"**{_tot_runs} prospecciones guardadas** para este cliente")
                _h1, _h2, _h3, _h4 = st.columns(4)
                _h1.metric("Total prospecciones", _tot_runs)
                _h2.metric("Total contactos",     _tot_leads)
                _h3.metric("Total con email",     _tot_emails,
                           f"{_tot_emails/_tot_leads*100:.0f}%" if _tot_leads else "")
                _h4.metric("Total con teléfono",  _tot_phones,
                           f"{_tot_phones/_tot_leads*100:.0f}%" if _tot_leads else "")

                st.divider()

                # ── Detalle por run ────────────────────────────────────────
                for _run7 in _runs7:
                    _st7   = _run7.get("stats") or {}
                    _fecha7 = _run7.get("run_date","")[:16].replace("T"," ")
                    _tot7  = _st7.get("total",0)
                    _em7   = _st7.get("emails",0)
                    _ph7   = _st7.get("phones",0)
                    _lu7   = _st7.get("lusha",0)

                    with st.expander(
                        f"📅 {_fecha7}  ·  **{_tot7} contactos**  ·  "
                        f"{_em7} emails  ·  {_ph7} teléfonos"
                        + (f"  ·  {_lu7} Lusha" if _lu7 else "")
                    ):
                        _ctcs7 = _run7.get("contactos",[])
                        if _ctcs7:
                            _rows7 = [{
                                "Nombre"   : norm_person(c.get("full_name",""), c.get("first_name",""), c.get("last_name","")),
                                "Empresa"  : norm_company(c.get("company_name","")),
                                "Cargo"    : c.get("job_title",""),
                                "Email"    : c.get("email",""),
                                "Teléfono" : c.get("phone",""),
                                "País"     : c.get("country",""),
                            } for c in _ctcs7]
                            st.dataframe(_pd.DataFrame(_rows7),
                                         use_container_width=True, hide_index=True)
                            try:
                                _particularidades7 = (st.session_state.selected_client or {}).get("particularidades_prospeccion","")
                                _xls7  = build_excel_hubspot(_ctcs7, _run7.get("empresas",[]), client_name, particularidades=_particularidades7)
                                _fname7 = (f"hubspot_{client_name.replace(' ','_')}_"
                                           f"{_fecha7.replace(' ','_').replace(':','')}.xlsx")
                                st.download_button(
                                    f"⬇️ Descargar Excel HubSpot ({_fecha7})",
                                    data=_xls7, file_name=_fname7,
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    key=f"dl7_{_run7['id']}",
                                )
                            except Exception as _e7:
                                st.error(f"Error generando Excel: {_e7}")
        except Exception as _e7x:
            st.error(f"Error cargando historial: {_e7x}")

# ── TAB 8 · EMPRESAS RECHAZADAS ───────────────────────────────────────────────
with tab8:
    st.subheader("🚫 Empresas Rechazadas")
    st.caption(
        "Registro de todas las empresas que rechazaste. "
        "La IA no volverá a recomendarlas mientras estén aquí. "
        "Puedes volver a aprobarlas o eliminarlas definitivamente."
    )

    # ── Verificar si las columnas necesarias existen en Supabase ──────────────
    _client_raw = st.session_state.selected_client or {}
    _faltan_cols = [
        col for col in ["empresas_rechazadas", "processed_domains", "export_history"]
        if col not in _client_raw
    ]
    if _faltan_cols:
        st.error(
            f"⚠️ **Columnas faltantes en Supabase**: `{'`, `'.join(_faltan_cols)}`\n\n"
            "Por eso los datos desaparecen al recargar — no se pueden guardar en la base de datos.\n\n"
            "**Solución:** Ve a [Supabase → SQL Editor](https://supabase.com/dashboard) y ejecuta:"
        )
        st.code(
            "ALTER TABLE clients ADD COLUMN IF NOT EXISTS empresas_rechazadas JSONB DEFAULT '[]'::jsonb;\n"
            "ALTER TABLE clients ADD COLUMN IF NOT EXISTS processed_domains    JSONB DEFAULT '[]'::jsonb;\n"
            "ALTER TABLE clients ADD COLUMN IF NOT EXISTS export_history       JSONB DEFAULT '[]'::jsonb;",
            language="sql"
        )
        st.info("Después de ejecutar el SQL, recarga la página y los datos se guardarán correctamente.")
        st.divider()

    _rechazadas = st.session_state.empresas_rechazadas or []

    if not _rechazadas:
        st.info("No hay empresas rechazadas aún. Cuando rechaces empresas en la pestaña **Empresas** y confirmes la selección, aparecerán aquí.")
    else:
        # ── Métricas ──────────────────────────────────────────────────────
        _con_razon = sum(1 for r in _rechazadas if r.get("razon_rechazo","").strip())
        rc1, rc2 = st.columns(2)
        rc1.metric("Total rechazadas", len(_rechazadas))
        rc2.metric("Con razón de rechazo", _con_razon,
                   help="Cuántas tienen una razón escrita — útil para entrenar las recomendaciones")

        st.divider()

        # ── Lista de rechazadas ────────────────────────────────────────────
        _to_reapprove = []  # dominios a re-aprobar
        _to_delete    = []  # dominios a eliminar permanentemente

        for _idx, _rec in enumerate(_rechazadas):
            _nom = _rec.get("nombre_empresa", "Sin nombre")
            _dom = _rec.get("dominio_web", "")
            _ind = _rec.get("industria", "")
            _pai = _rec.get("pais", "")
            _rf  = _rec.get("razon_fit", "")
            _rr  = _rec.get("razon_rechazo", "")
            _fec = _rec.get("fecha_rechazo", "")

            with st.container():
                col_info, col_razon, col_acciones = st.columns([3, 4, 2])

                with col_info:
                    st.markdown(f"**{_nom}**")
                    _meta = " · ".join(filter(None, [_dom, _ind, _pai, _fec]))
                    st.caption(_meta)
                    if _rf:
                        with st.expander("Ver razón de fit original", expanded=False):
                            st.caption(_rf)

                with col_razon:
                    _nueva_razon = st.text_area(
                        "Razón del rechazo (opcional)",
                        value=_rr,
                        height=80,
                        placeholder="Ej: precio muy alto, no tienen equipo de ventas, ya tienen proveedor...",
                        key=f"razon_rechazo_{_idx}",
                        label_visibility="collapsed",
                    )
                    if _nueva_razon != _rr:
                        st.session_state.empresas_rechazadas[_idx]["razon_rechazo"] = _nueva_razon

                with col_acciones:
                    st.write("")
                    if st.button("✅ Re-aprobar", key=f"reap_{_idx}", use_container_width=True,
                                 help="Quitar de rechazadas y agregar a la lista de empresas para prospectar"):
                        _to_reapprove.append(_dom)
                    if st.button("🗑️ Eliminar", key=f"del_{_idx}", use_container_width=True,
                                 help="Eliminar de este registro (la IA podrá volver a recomendarla)"):
                        _to_delete.append(_dom)

            st.markdown("---")

        # ── Procesar acciones ──────────────────────────────────────────────
        _changed = False

        if _to_reapprove:
            for _dom_ra in _to_reapprove:
                # Encontrar la empresa rechazada y volver a agregarla como aprobada
                _emp_ra = next((r for r in st.session_state.empresas_rechazadas if r.get("dominio_web") == _dom_ra), None)
                if _emp_ra:
                    _emp_nuevo = {
                        "nombre_empresa"  : _emp_ra.get("nombre_empresa",""),
                        "dominio_web"     : _dom_ra,
                        "industria"       : _emp_ra.get("industria",""),
                        "pais"            : _emp_ra.get("pais",""),
                        "razon_fit"       : _emp_ra.get("razon_fit",""),
                        "linkedin_url"    : _emp_ra.get("linkedin_url",""),
                        "tamano_empleados": _emp_ra.get("tamano_empleados",""),
                        "aprobada"        : True,
                    }
                    st.session_state.empresas.append(_emp_nuevo)
                    st.session_state.empresas_aprobadas.append(_emp_nuevo)
                # Quitar de rechazadas
                st.session_state.empresas_rechazadas = [
                    r for r in st.session_state.empresas_rechazadas
                    if r.get("dominio_web") != _dom_ra
                ]
            _changed = True

        if _to_delete:
            st.session_state.empresas_rechazadas = [
                r for r in st.session_state.empresas_rechazadas
                if r.get("dominio_web") not in _to_delete
            ]
            _changed = True

        # Guardar razones de rechazo actualizadas en Supabase
        if st.button("💾 Guardar razones de rechazo", key="save_razones_rechazo", type="primary"):
            _db8 = get_db()
            if _db8 and st.session_state.selected_client_id:
                try:
                    _db8.update_client(st.session_state.selected_client_id, {
                        "empresas_rechazadas": st.session_state.empresas_rechazadas
                    })
                    st.success("✅ Razones guardadas correctamente.")
                except Exception as _e8:
                    if "400" in str(_e8) or "column" in str(_e8).lower():
                        st.error(
                            "❌ La columna `empresas_rechazadas` no existe en Supabase. "
                            "Ve a Supabase → SQL Editor y ejecuta:\n\n"
                            "```sql\n"
                            "ALTER TABLE clients ADD COLUMN IF NOT EXISTS empresas_rechazadas JSONB DEFAULT '[]'::jsonb;\n"
                            "ALTER TABLE clients ADD COLUMN IF NOT EXISTS processed_domains JSONB DEFAULT '[]'::jsonb;\n"
                            "```"
                        )
                    else:
                        st.error(f"Error al guardar: {_e8}")
            else:
                st.warning("Supabase no configurado.")

        if _changed:
            # Persistir cambios en Supabase
            _db8c = get_db()
            if _db8c and st.session_state.selected_client_id:
                try:
                    _db8c.update_client(st.session_state.selected_client_id, {
                        "empresas_rechazadas": st.session_state.empresas_rechazadas
                    })
                except Exception: pass
            st.rerun()

        # ── Resumen de razones frecuentes (para entrenar IA) ──────────────
        _razones_texto = [r.get("razon_rechazo","").strip() for r in _rechazadas if r.get("razon_rechazo","").strip()]
        if _razones_texto:
            st.divider()
            with st.expander("📊 Resumen de razones de rechazo (para entrenar la IA)", expanded=False):
                st.caption(
                    "Estas razones se incluirán automáticamente como contexto en la próxima "
                    "recomendación de empresas de la IA, para que aprenda qué tipo de empresa NO recomendar."
                )
                for _rz in _razones_texto:
                    st.markdown(f"- {_rz}")

